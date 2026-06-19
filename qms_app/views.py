import os
from urllib import request
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils import timezone
from datetime import datetime
from django.db import models
from django import forms
from decimal import Decimal, InvalidOperation
from django.conf import settings
from django.http import Http404
from django.http import HttpResponseForbidden
from uuid import UUID
import uuid
from form_builder.services import build_traceability_data
from django.urls import reverse
from difflib import HtmlDiff
from datetime import date, datetime
from django.utils.html import escape
from difflib import HtmlDiff
from .models import ( Form, Stage, SubStage, FormField )
from .forms import RegisterForm, LoginForm,generate_form
from .models import ( CustomUser, STAGES, STAGE_ROLE_MAP )
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.contrib.auth import authenticate
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import SignatureVerification
from .models import QMSDocument, QMSDocumentVersion,Form, Stage, FormSubmission,FormField,DocumentRevision,SubStage,FormFolder, BatchPart
from .utils.docx_to_html import convert_docx_to_html
from .utils.html_to_docx import convert_html_to_docx
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
# from .utils.docx_to_html import convert_docx_to_html_with_pages
import json
from django.contrib.auth import get_user_model
from .models import UserPageAccess, FormFolder, NCR
import random
import string
from django.contrib.auth.hashers import check_password
from django.db.models import Max
from .decorators import require_page_permission
from .models import QMSDocument, DocumentFolder
from audit_log.models import AuditLog
from django.http import JsonResponse
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.utils import timezone
from datetime import date, datetime
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
import difflib
from bs4 import BeautifulSoup
from student_test.models import StudentResult
from audit_log.models import AuditLog
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from .models import (
    BatchPart,
    Form,
    QMSDocument,
    DocumentRevision,
    SignatureVerification,
    StageHistory,
    CAPA,
    CustomUser,
    MaterialBatch,
    DocumentWorkflowLog,
    Certificate,
    CertificateCategory
)
from datetime import timedelta
from django.db.models import Max
from django.db import transaction
from django.views.decorators.http import require_POST
from po_qu.models import RFQAttachment
from django.contrib.auth import update_session_auth_hash



# ==============================================================================================
# ================================== Authentication Views ======================================
# ==============================================================================================
@login_required
def register_view(request):

    if not request.user.is_superuser:
        return redirect("denied")

    if request.method == 'POST':
        form = RegisterForm(request.POST)

        if form.is_valid():

            password = form.cleaned_data['password']

            try:
                validate_password(password)
            except ValidationError as e:
                form.add_error("password", e)
                return render(request, 'qms_app/register.html', {'form': form})

            user = form.save(commit=False)
            user.set_password(password)


            user.password_changed_at = timezone.now()

            user.save()


            if user.role != "student":
                UserPageAccess.objects.get_or_create(user=user)

            return redirect('page_settings')

    else:
        form = RegisterForm()

    return render(request, 'qms_app/register.html', {
        'form': form
    })


# ==================================================================================================
# ======================================= Login / Logout views =====================================
# ==================================================================================================

def login_view(request):

    if request.user.is_authenticated:
        if request.user.role == "student":
            return redirect("test_list")
        return redirect("main_dashboard")

    if request.method == 'POST':

        form = LoginForm(request=request, data=request.POST)

        email = request.POST.get("username")

        try:
            user = CustomUser.objects.get(email=email)
        except CustomUser.DoesNotExist:
            user = None

      
        if user and user.account_locked:
            form.add_error(None, "Account locked after 3 failed attempts. Contact Admin.")
            return render(request, 'qms_app/login.html', {'form': form})

        if form.is_valid():

            user = form.get_user()

      
            # LOGIN USER FIRST
            login(request, user)

            # PASSWORD EXPIRY CHECK
            if user.password_changed_at:
                if user.password_changed_at + timedelta(days=60) < timezone.now():
                    return redirect("change_password")

            # RESET FAILED ATTEMPTS
            user.failed_attempts = 0
            user.save()

   
            AuditLog.objects.create(
                user=user,
                role=user.role,
                module="Authentication",
                action="LOGIN",
                model_name="CustomUser",
                object_repr=user.email,
                new_data={"message": "User logged in"},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            if user.role == "student":
                return redirect("test_list")

            return redirect("main_dashboard")

        else:
            
            if user:
                user.failed_attempts += 1

                if user.failed_attempts >= 3:
                    user.account_locked = True

                user.save()

    else:
        form = LoginForm()

    return render(request, 'qms_app/login.html', {'form': form})





@login_required
def change_password(request):

    if request.method == "POST":

        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        # CHECK OLD PASSWORD
        if not request.user.check_password(old_password):

            messages.error(
                request,
                "Current password is incorrect"
            )

            return redirect("change_password")

        # CHECK PASSWORD MATCH
        if new_password != confirm_password:

            messages.error(
                request,
                "Passwords do not match"
            )

            return redirect("change_password")

        # OPTIONAL PASSWORD VALIDATION
        try:
            validate_password(new_password, request.user)

        except ValidationError as e:

            for error in e.messages:
                messages.error(request, error)

            return redirect("change_password")

        # SET NEW PASSWORD
        request.user.set_password(new_password)

        # UPDATE PASSWORD CHANGE TIME
        request.user.password_changed_at = timezone.now()

        # SAVE USER
        request.user.save()

        # KEEP USER LOGGED IN
        update_session_auth_hash(
            request,
            request.user
        )

        messages.success(
            request,
            "Password updated successfully"
        )

        return redirect("main_dashboard")

    return render(
        request,
        "qms_app/change_password.html"
    )


    
@login_required
def logout_view(request):

    # AUDIT LOG
    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="Authentication",
        action="LOGOUT",
        model_name="CustomUser",
        object_repr=request.user.email,
        old_data={"message": "User logged out"},
        ip_address=request.META.get("REMOTE_ADDR")
    )

    logout(request)

    return redirect('login')


def denied(request):
    return render(request, 'qms_app/denied.html')

# =======================================================================================================
# ========================================== Access Control =============================================
# =======================================================================================================

def user_has_stage_access(user, process):
    if user.is_superuser or user.role == "admin":
        return True
    if user == process.created_by:
        return True

    return user.role == STAGE_ROLE_MAP.get(process.stage)



# ========================================================================================================
# =========================================== Dashboard ==================================================
# ========================================================================================================
@require_page_permission("can_dashboard")
@login_required
def dashboard(request):

    if request.user.role == "student":
        return redirect("test_list")

    # =====================================================
    #  PRODUCTION METRICS
    # =====================================================

    active_parts = BatchPart.objects.filter(is_active=True).count()

    completed_parts = 0
    forms = Form.objects.prefetch_related("stages")

    for form in forms:
        last_stage = form.stages.order_by("-order").first()
        if last_stage:
            completed_parts += BatchPart.objects.filter(
                current_stage=last_stage
            ).count()

    # =====================================================
    #  DOCUMENT CONTROL
    # =====================================================

    total_documents = QMSDocument.objects.count()

    this_month = timezone.now().replace(day=1)

    revised_this_month = DocumentRevision.objects.filter(
        edited_at__gte=this_month
    ).count()

    # =====================================================
    #  APPROVAL CONTROL
    # =====================================================

    prepared_count = SignatureVerification.objects.filter(slot="prepared").count()
    approved_count = SignatureVerification.objects.filter(slot="approved").count()
    pending_approvals = max(prepared_count - approved_count, 0)

    # =====================================================
    #  TRAINING COMPLIANCE
    # =====================================================

    total_students = CustomUser.objects.filter(role="student").count()

    passed_students = StudentResult.objects.filter(
        status="PASS"
    ).values("student").distinct().count()

    training_compliance = 0
    if total_students > 0:
        training_compliance = round(
            (passed_students / total_students) * 100, 2
        )

    # =====================================================
    #  CAPA METRICS
    # =====================================================

    total_capa = CAPA.objects.count()
    open_capa = CAPA.objects.filter(status="OPEN").count()

    in_progress_capa = CAPA.objects.filter(
        status__in=["INVESTIGATION", "ACTION", "VERIFICATION"]
    ).count()

    closed_capa = CAPA.objects.filter(status="CLOSED").count()

    overdue_capa = CAPA.objects.filter(
        due_date__lt=timezone.now().date(),
        status__in=["OPEN", "INVESTIGATION", "ACTION"]
    ).count()

    # =====================================================
    #  MONTHLY PRODUCTION TREND
    # =====================================================

    monthly_data = (
        StageHistory.objects
        .annotate(month=TruncMonth("moved_at"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )

    months = []
    counts = []

    for item in monthly_data:
        if item["month"]:
            months.append(item["month"].strftime("%b %Y"))
            counts.append(item["total"])

    # =====================================================
    #  RECENT ACTIVITY
    # =====================================================

    recent_logs = AuditLog.objects.select_related("user")[:8]

    # =====================================================
    # EXECUTIVE SEARCH (SAFE VERSION)
    # =====================================================

    search_result = None

    user_email = request.GET.get("user")
    batch_id = request.GET.get("batch_id")
    part_id = request.GET.get("part_id")
    company =request.GET.get("company")

    if user_email or batch_id or part_id:

        parts = BatchPart.objects.select_related(
            "batch",
            "batch__form",
            "current_stage",
            "created_by"
        )

        if user_email:
            parts = parts.filter(created_by__email=user_email)

        if batch_id:
            parts = parts.filter(batch__batch_id=batch_id)

        if part_id:
            parts = parts.filter(part_id=part_id)

        if company:
            parts = parts.fillter(batch__company_name__icontains=company)

        part = parts.first()

        if part:
            search_result = {
                "company_name":part.batch.company_name,
                "folder_name": (
                    part.batch.form.folder.name
                    if part.batch and part.batch.form and part.batch.form.folder
                    else "N/A"
                ),
                "form_name": (
                    part.batch.form.name
                    if part.batch and part.batch.form
                    else "N/A"
                ),
                "stage_name": (
                    part.current_stage.name
                    if part.current_stage
                    else "N/A"
                ),
                "status": (
                    "Active" if part.is_active else "Closed"
                ),
                "created_by": (
                    part.created_by.email
                    if part.created_by
                    else "N/A"
                ),
                "created_at": part.created_at,
                "batch_id": part.batch.batch_id,
                "part_id": part.part_id,
            }

    # =====================================================
    # ------ RENDER ------
    # =====================================================

    return render(
        request,
        "qms_app/dashboard.html",
        {
            "active_parts": active_parts,
            "completed_parts": completed_parts,
            "total_documents": total_documents,
            "revised_this_month": revised_this_month,
            "pending_approvals": pending_approvals,
            "training_compliance": training_compliance,
            "total_capa": total_capa,
            "open_capa": open_capa,
            "in_progress_capa": in_progress_capa,
            "closed_capa": closed_capa,
            "overdue_capa": overdue_capa,
            "months": json.dumps(months),
            "counts": json.dumps(counts),
            "recent_logs": recent_logs,
            "search_result": search_result,
        }
    )

# ====================================================================================================================
# ================================================ SOP ===============================================================
# ====================================================================================================================

@require_page_permission("can_sop")
def sop(request):
    return render(request, 'qms_app/sop/sop.html')

# =====================================================================================================================
# ================================================ Document upload ====================================================
# =====================================================================================================================


@login_required
def upload_document(request):

    if request.method == "POST":

        # ================= BASIC DATA =================
        title = request.POST.get("title")
        clause = request.POST.get("clause")
        file = request.FILES.get("file")

        # ✅ MANY TO MANY (IMPORTANT)
        certificate_ids = request.POST.getlist("certificate")
        category_ids = request.POST.getlist("certificate_category")

        # ================= VALIDATION =================
        if not title:
            return JsonResponse({"error": "Title is required"}, status=400)

        if not file:
            return JsonResponse({"error": "File is required"}, status=400)

        # ================= FOLDER =================
        folder_id = request.POST.get("folder_id")
        new_folder = request.POST.get("new_folder")

        folder = None

        if new_folder:
            folder = DocumentFolder.objects.create(
                name=new_folder,
                created_by=request.user
            )

        elif folder_id:
            folder = get_object_or_404(DocumentFolder, id=folder_id)

        # ================= CREATE DOCUMENT =================
        doc = QMSDocument.objects.create(
            title=title,
            clause=clause,
            original_file=file,
            created_by=request.user,
            folder=folder,
            target_folder=folder,
            assigned_to=request.user
        )

        # ================= SET MANY TO MANY =================
        doc.certificate.set(certificate_ids)
        doc.certificate_category.set(category_ids)

        # ================= DEBUG (REMOVE LATER) =================
        print("CERT IDS:", certificate_ids)
        print("CATEGORY IDS:", category_ids)

        # ================= AUDIT =================
        AuditLog.objects.create(
            user=request.user,
            role=request.user.role,
            module="Document",
            action="CREATE",
            model_name="QMSDocument",
            object_repr=doc.title,
            new_data={
                "folder": folder.name if folder else None,
                "certificate_ids": certificate_ids,
                "category_ids": category_ids,
                "clause": clause,
            },
            ip_address=request.META.get("REMOTE_ADDR")
        )

        return redirect("edit_document", doc.id)

    # ================= GET (PAGE LOAD) =================
    folders = DocumentFolder.objects.all()
    certificates = Certificate.objects.all()
    categories = CertificateCategory.objects.all()

    return render(request, "qms_app/sop/upload_document.html", {
        "folders": folders,
        "certificates": certificates,
        "categories": categories
    })




# @require_page_permission("can_documents")
# @login_required
# def document_home(request, folder_id=None):

#     # ✅ ALL USERS SEE EVERYTHING
#     folders = DocumentFolder.objects.all()

#     if folder_id:
#         current_folder = get_object_or_404(DocumentFolder, id=folder_id)
#         documents = QMSDocument.objects.filter(folder=current_folder,status='completed')
#     else:
#         current_folder = None
#         documents = []

#     return render(request, 'qms_app/sop/document_home.html', {
#         'documents': documents,
#         'folders': folders,
#         'current_folder': current_folder
#     })


#     #  Get or create page access safely
#     page_access, created = UserPageAccess.objects.get_or_create(
#         user=request.user
#     )

#     #  Page restriction
#     if not page_access.can_documents:
#         return redirect("denied")

#     #  Folder restriction (if you want same folder logic)
#     allowed_folders = page_access.allowed_folders.all()

#     if folder_id:
#         current_folder = allowed_folders.filter(id=folder_id).first()

#         if not current_folder:
#             return redirect("denied")

#         documents = QMSDocument.objects.filter(folder=current_folder,status='completed')
#     else:
#         current_folder = None
#         documents = []

#     return render(request, 'qms_app/sop/document_home.html', {
#         'documents': documents,
#         'folders': allowed_folders,
#         'current_folder': current_folder
#     })
# @require_page_permission("can_documents")
# @login_required
# def document_home(request, folder_id=None):

#     # ✅ Get access FIRST (for all users)
#     page_access, created = UserPageAccess.objects.get_or_create(
#         user=request.user
#     )

#     # ✅ Superuser bypass
#     if request.user.is_superuser:

#         folders = DocumentFolder.objects.all()

#         if folder_id:
#             current_folder = get_object_or_404(DocumentFolder, id=folder_id)
#             documents = QMSDocument.objects.filter(
#                 folder=current_folder,
#                 status="completed"
#             )
#         else:
#             current_folder = None
#             documents = []

#         return render(request, 'qms_app/sop/document_home.html', {
#             'documents': documents,
#             'folders': folders,
#             'current_folder': current_folder,
#             'access': page_access   # 🔥 ADD THIS
#         })

#     # ✅ Page restriction
#     if not page_access.can_documents:
#         return redirect("denied")

#     # ✅ Folder restriction
#     allowed_folders = page_access.allowed_folders.all()

#     if folder_id:
#         current_folder = allowed_folders.filter(id=folder_id).first()

#         if not current_folder:
#             return redirect("denied")

#         documents = QMSDocument.objects.filter(
#             folder=current_folder,
#             status="completed"
#         )
#     else:
#         current_folder = None
#         documents = []

#     return render(request, 'qms_app/sop/document_home.html', {
#         'documents': documents,
#         'folders': allowed_folders,
#         'current_folder': current_folder,
#         'access': page_access   # 🔥 ADD THIS
#     })
@require_page_permission("can_documents")
@login_required
def document_home(request, folder_id=None):

    # Get access
    page_access, created = UserPageAccess.objects.get_or_create(
        user=request.user
    )

    # Page restriction only
    if not page_access.can_documents:
        return redirect("denied")

    # ✅ SHOW ALL DOCUMENT FOLDERS (NO RESTRICTION)
    folders = DocumentFolder.objects.all()

    if folder_id:
        current_folder = get_object_or_404(DocumentFolder, id=folder_id)

        documents = QMSDocument.objects.filter(
            folder=current_folder,
            status="completed"
        )
    else:
        current_folder = None
        documents = []

    return render(request, 'qms_app/sop/document_home.html', {
        'documents': documents,
        'folders': folders,
        'current_folder': current_folder
    })

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import DocumentFolder, QMSDocument, Certificate

@login_required
def folder_documents(request, folder_id):

    folder = get_object_or_404(DocumentFolder, id=folder_id)

    documents = QMSDocument.objects.filter(
        folder=folder,
        status='completed'
    ).prefetch_related(
        'versions',
        'certificate_category__certificate'   # ✅ IMPORTANT (for performance)
    )

    for doc in documents:
        latest = doc.versions.order_by('-version_number').first()
        doc.latest_version = latest.version_number if latest else 1

   
    certificates = Certificate.objects.all()

    return render(request, 'qms_app/sop/folder_documents.html', {
        'folder': folder,
        'documents': documents,
        'certificates': certificates   # ✅ PASS TO TEMPLATE
    })



# =============================================================================================================
# ============================================ Folder Creation ==============================================
# =============================================================================================================

@login_required
def create_folder(request):
    if request.method == "POST":
        name = request.POST.get('name')

        if name:
            DocumentFolder.objects.create(
                name=name,
                created_by=request.user
            )

    return redirect('document_home')


#  DELETE FOLDER
@login_required
def delete_folder(request, id):

    folder = get_object_or_404(DocumentFolder, id=id)

    # Optional safety
    if QMSDocument.objects.filter(folder=folder).exists():
        return redirect('document_home')

    folder.delete()

    return redirect('document_home')




@login_required
def view_document(request, doc_id):

    doc = get_object_or_404(QMSDocument, id=doc_id)

    # Detect file extension
    file_ext = os.path.splitext(doc.original_file.name)[1].lower()

    # ======================= PDF FLOW =========================
    if file_ext == ".pdf":
        return render(request, "qms_app/sop/view_document.html", {
            "doc": doc,
            "is_pdf": True,
            "html_pages": None,
            "revisions": None,
            "pdf_signatures": doc.pdf_signatures or "[]",
        })

    # ======================= DOCX FLOW ========================
    elif file_ext == ".docx":

        latest_revision = doc.revisions.order_by('-edited_at').first()

        # ✅ FIXED PAGE SPLIT
        if latest_revision and latest_revision.edited_html:

            html = latest_revision.edited_html

            # 🔥 USE SAME SPLIT AS EDITOR
            page_divs = html.split('<div class="page-break"></div>')

            html_pages = [
                p.strip() for p in page_divs if p.strip()
            ]

        else:
            try:
                html_pages = convert_docx_to_html(doc.original_file.path)
            except Exception as e:
                html_pages = [f"<p>Error loading document: {str(e)}</p>"]

        # ✅ GET ALL REVISIONS
        revisions = doc.revisions.order_by('-edited_at')

        return render(request, "qms_app/sop/view_document.html", {
            "doc": doc,
            "html_pages": html_pages,
            "revisions": revisions,
            "is_pdf": False
        })

    # =================== UNSUPPORTED FILE =====================
    else:
        return render(request, "qms_app/sop/view_document.html", {
            "doc": doc,
            "unsupported": True,
            "html_pages": None,
            "revisions": None
        })




from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required

@login_required
def render_signed_pdf(request, doc_id):
    doc = get_object_or_404(QMSDocument, id=doc_id)

    if not doc.pdf_signatures:
        return redirect(doc.original_file.url)
    reader = PdfReader(doc.original_file.path)
    writer = PdfWriter()

    for page_num, page in enumerate(reader.pages):
        packet = BytesIO()
        can = canvas.Canvas(packet)

        has_signature = False

        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)

        for sig in doc.pdf_signatures:
            if int (sig.get("page", 0)) == page_num:
                x = sig.get("x", 0.5) * page_width
                y = sig.get("y", 0.5) * page_height

                y = page_height - y

                raw = sig.get("html", "")
                text = re.sub('<[^<]+?>', '', raw)
                can.setFont("Helvetica", 10)
                can.drawString(x, y, text)

                has_signature = True

        can.save()
        packet.seek(0)

        overlay = PdfReader(packet)

        if has_signature and len(overlay.pages) > 0:
            page.merge_page(overlay.pages[0])
        writer.add_page(page)

    output = BytesIO()
    writer.write(output)
    output.seek(0)

    return HttpResponse(output, content_type="application/pdf")

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

@login_required
def delete_documents(request, doc_id):
    if request.method == "POST":
        doc = get_object_or_404(QMSDocument, id=doc_id)
        doc.delete()

        return JsonResponse({
            "status": "success",
            "message": "Document deleted"
        })

    return JsonResponse({"status": "error"}, status=400)

# ========================================================================================
# ================================= Edit Document ========================================
# ========================================================================================
# @login_required
# def edit_document(request, doc_id):

#     doc = get_object_or_404(QMSDocument, id=doc_id)

#     page_access, _ = UserPageAccess.objects.get_or_create(user=request.user)

#     # Detect file extension
#     file_ext = os.path.splitext(doc.original_file.name)[1].lower()


#     # ================= PDF → NOT EDITABLE ====================
#     if file_ext == ".pdf":
#         # PDF should not be edited → redirect to view page
#         return redirect("view_document", doc_id=doc.id)


#     # ================= DOCX EDIT FLOW ========================
#     # Fetch the latest saved version
#     latest_version = doc.versions.order_by('-version_number').first()

#     if latest_version and latest_version.edited_html:
#         html = latest_version.edited_html

#         # Clean HTML safely
#         soup = BeautifulSoup(html, 'html.parser')

#         # Split pages
#         page_divs = html.split('<div class="page-break"></div>')
#         html_pages = [p.strip() for p in page_divs if p.strip()]

#     else:
#         # First-time document → convert from DOCX
#         html_pages = convert_docx_to_html(doc.original_file.path)

#     return render(request, "qms_app/sop/edit_document.html", {
#         "doc": doc,
#         "html_pages": html_pages,
#         "access": page_access
#     })

from openpyxl import load_workbook

@login_required
def edit_document(request, doc_id):

    doc = get_object_or_404(QMSDocument, id=doc_id)
    access = UserPageAccess.objects.filter(user=request.user).first()

    file_ext = os.path.splitext(doc.original_file.name)[1].lower()

    # ================= PDF → VIEW ONLY ====================
    if file_ext == ".pdf":
        return render(request, "qms_app/sop/edit_document.html", {
            "doc": doc,
            "html_pages": [],
            "is_excel": False,
            "is_image": False,
            "is_pdf": True   # ✅ NEW FLAG
        })
    # ================= IMAGE FILE ====================
    if file_ext in [".jpg", ".jpeg", ".png", ".webp"]:

        return render(request, "qms_app/sop/edit_document.html", {
            "doc": doc,
            "html_pages": [],
            "is_excel": False,
            "is_image": True   # ✅ IMPORTANT FLAG
        })
    # ================= EXCEL FILE ====================
    if file_ext in [".xlsx", ".xls"]:

        excel_data = []

        # ✅ If already saved → use DB
        if doc.excel_data:
            excel_data = doc.excel_data

        else:
            try:
                wb = load_workbook(doc.original_file.path)
                sheet = wb.active

                # 🔥 FIX: Convert None → ""
                for row in sheet.iter_rows(values_only=True):
                    clean_row = [cell if cell is not None else "" for cell in row]
                    excel_data.append(clean_row)

            except Exception as e:
                print("EXCEL READ ERROR:", str(e))

        return render(request, "qms_app/sop/edit_document.html", {
            "doc": doc,
            "html_pages": [],
            "is_excel": True,
            "excel_data": excel_data
        })

    # ================= DOCX EDIT FLOW ========================
    # ================= DOCX EDIT FLOW ========================

    html_pages = []
    latest_revision = doc.revisions.order_by('-edited_at').first()

    if doc.status != "completed" and latest_revision and latest_revision.edited_html:

        html = latest_revision.edited_html

        page_divs = html.split('<div class="page-break"></div>')
        html_pages = [p.strip() for p in page_divs if p.strip()]


    else:
        latest_version = doc.versions.order_by('-version_number').first()

        if latest_version and latest_version.edited_html:

            html = latest_version.edited_html

            page_divs = html.split('<div class="page-break"></div>')
            html_pages = [p.strip() for p in page_divs if p.strip()]

        else:
            
            try:
    # 🔥 CHECK FILE EXISTS
                if not doc.original_file or not os.path.exists(doc.original_file.path):
                    raise Exception("File not found on server")

                print("FILE PATH:", doc.original_file.path)

                html_pages = convert_docx_to_html(doc.original_file.path)

            except Exception as e:
                import traceback
                print("DOCX ERROR FULL:")
                traceback.print_exc()

                return render(request, "qms_app/sop/error.html", {
                    "message": f"DOCX conversion failed: {str(e)}"
                })

    return render(request, "qms_app/sop/edit_document.html", {
        "doc": doc,
        "html_pages": html_pages,
        "is_excel": False
    })


@login_required
def revise_document(request, doc_id):

    doc = get_object_or_404(QMSDocument, id=doc_id)

    if doc.status != "completed":
        return JsonResponse({"error": "Only completed docs can be revised"}, status=400)

    from bs4 import BeautifulSoup

    latest_version = doc.versions.order_by('-version_number').first()

    if latest_version:
        soup = BeautifulSoup(latest_version.edited_html, "html.parser")

        # ❌ remove all signatures
        for sig in soup.select(".signature-box"):
            sig.decompose()

        clean_html = str(soup)

        # ✅ save as new draft
        DocumentRevision.objects.create(
            document=doc,
            edited_by=request.user,
            edited_html=clean_html,
            change_summary="Revision started (signatures removed)"
        )

    # 🔄 reset workflow
    doc.status = "process_owner"
    doc.assigned_to = doc.created_by
    doc.last_message = "Document sent for revision"
    doc.is_read = False
    doc.save()

    # 📝 log history
    DocumentWorkflowLog.objects.create(
        document=doc,
        stage="process_owner",
        action="Revision Restarted",
        user=request.user,
        message="New revision cycle started"
    )

    return JsonResponse({"status": "ok"})



# ========================================================================================
# ================================= Save Document ========================================
# ========================================================================================
# @login_required
# def save_document(request, doc_id):
#     if request.method != "POST":
#         return JsonResponse(
#             {"status": "error", "message": "Invalid request method"}, 
#             status=405
#         )

#     doc = get_object_or_404(QMSDocument, id=doc_id)

#     # ----- Read posted JSON -----
#     try:
#         data = json.loads(request.body.decode("utf-8"))
#         edited_html = data.get("content", "").strip()
#         change_summary = data.get("summary", "").strip()

#         new_title = data.get("title", "").strip()
#     except Exception as e:
#         return JsonResponse(
#             {"status": "error", "message": f"Invalid JSON: {e}"},
#             status=400
#         ) 

#     if not edited_html:
#         return JsonResponse({"status": "error", "message": "Content cannot be empty"}, status=400)

#     # ----- Get latest revision -----
#     latest_revision = doc.revisions.order_by("-edited_at").first()
#     old_html = latest_revision.edited_html if latest_revision else ""

#     # ----- Strip HTML safely using BeautifulSoup -----
#     def plain_text(html):
#         return BeautifulSoup(html or "", "html.parser").get_text(separator=" ")

#     old_text = plain_text(old_html)
#     new_text = plain_text(edited_html)

#     # ----- AUTO DIFF (word-level) -----
#     if not change_summary:
#         old_words = old_text.split()
#         new_words = new_text.split()

#         diff_summary = []
#         sm = difflib.SequenceMatcher(a=old_words, b=new_words)

#         for tag, a0, a1, b0, b1 in sm.get_opcodes():
#             if tag == "replace":
#                 diff_summary.append(
#                     f"{' '.join(old_words[a0:a1])} → {' '.join(new_words[b0:b1])}"
#                 )
#             elif tag == "delete":
#                 diff_summary.append(
#                     f"Removed: {' '.join(old_words[a0:a1])}"
#                 )
#             elif tag == "insert":
#                 diff_summary.append(
#                     f"Added: {' '.join(new_words[b0:b1])}"
#                 )

#         change_summary = "\n".join(diff_summary) if diff_summary else "No textual changes detected."

#     # ----- Version number -----
#     version_number = doc.revisions.count() + 1
#     # version_number = None
#     # if doc.status == "completed":
#     #     last_version = doc.version.order_by('-version_number').first()
#     #     version_number = (last_version.version_number + 1) if last_version else 1

#     # ----- Save DOCX file -----
#     folder = os.path.join(settings.MEDIA_ROOT, "qms_docs", "versions")

 
#     if not os.path.exists(folder):
#         os.makedirs(folder, exist_ok=True)

#     output_filename = f"{doc.id}_v{version_number}.docx"
#     output_path = os.path.join(folder, output_filename)


#     try:
#         convert_html_to_docx(edited_html, output_path)
#     except Exception as e:
#         return JsonResponse({
#         "status": "error",
#         "message": f"DOCX generation failed: {str(e)}"
#     }, status=500)

#     # Save version entry
#     version = QMSDocumentVersion.objects.create(
#         document=doc,
#         version_number=version_number,
#         edited_by=request.user,
#         edited_html=edited_html,
#         edited_docx=f"qms_docs/versions/{output_filename}",
#     )
#     AuditLog.objects.create(
#         user=request.user,
#         role=request.user.role,
#         module="Documents",
#         action="UPDATE",
#         model_name="QMSDocument",
#         object_repr=doc.title,
#         new_data={"version": version_number},
#         ip_address=request.META.get("REMOTE_ADDR")
#     )


#     # Save revision entry
#     DocumentRevision.objects.create(
#         document=doc,
#         edited_by=request.user,
#         edited_html=edited_html,
#         change_summary=change_summary,
#         version_number=version_number,
#     )

#     # doc.status = "process_owner"
#     # doc.folder = None
   
#     if new_title and new_title !=doc.title:
#         doc.title = new_title
#     doc.save()
#     return JsonResponse({"status": "success", "version": version_number})

# @login_required
# def save_document(request, doc_id):

#     if request.method != "POST":
#         return JsonResponse(
#             {"status": "error", "message": "Invalid request method"}, 
#             status=405
#         )

#     doc = get_object_or_404(QMSDocument, id=doc_id)

#     # ================= READ DATA =================
#     try:
#         data = json.loads(request.body.decode("utf-8"))
#         edited_html = data.get("content", "").strip()
#         change_summary = data.get("summary", "").strip()
#         new_title = data.get("title", "").strip()
#     except Exception as e:
#         return JsonResponse(
#             {"status": "error", "message": f"Invalid JSON: {e}"},
#             status=400
#         )

#     if not edited_html:
#         return JsonResponse({
#             "status": "error",
#             "message": "Content cannot be empty"
#         }, status=400)

#     # ==========================================================
#     # 🔥 IMPORTANT FIX: BLOCK VERSION DURING WORKFLOW
#     # ==========================================================
#     if doc.status != "completed":

#         # 👉 Just update latest revision (NO new version)
#         latest_revision = doc.revisions.order_by("-edited_at").first()

#         if latest_revision:
#             latest_revision.edited_html = edited_html
#             latest_revision.save()
#         else:
#             # First time draft
#             DocumentRevision.objects.create(
#                 document=doc,
#                 edited_by=request.user,
#                 edited_html=edited_html,
#                 change_summary="Draft saved (no version)"
#             )

#         # Update title if needed
#         if new_title and new_title != doc.title:
#             doc.title = new_title
#             doc.save()

#         return JsonResponse({
#             "status": "success",
#             "message": "Draft saved (no version created)"
#         })

#     # ==========================================================
#     # ✅ VERSION CREATION (ONLY AFTER COMPLETED)
#     # ==========================================================

#     # ----- Get latest revision -----
#     latest_revision = doc.revisions.order_by("-edited_at").first()
#     old_html = latest_revision.edited_html if latest_revision else ""

#     def plain_text(html):
#         return BeautifulSoup(html or "", "html.parser").get_text(separator=" ")

#     old_text = plain_text(old_html)
#     new_text = plain_text(edited_html)

#     # ----- AUTO DIFF -----
#     if not change_summary:
#         old_words = old_text.split()
#         new_words = new_text.split()

#         diff_summary = []
#         sm = difflib.SequenceMatcher(a=old_words, b=new_words)

#         for tag, a0, a1, b0, b1 in sm.get_opcodes():
#             if tag == "replace":
#                 diff_summary.append(
#                     f"{' '.join(old_words[a0:a1])} → {' '.join(new_words[b0:b1])}"
#                 )
#             elif tag == "delete":
#                 diff_summary.append(
#                     f"Removed: {' '.join(old_words[a0:a1])}"
#                 )
#             elif tag == "insert":
#                 diff_summary.append(
#                     f"Added: {' '.join(new_words[b0:b1])}"
#                 )

#         change_summary = "\n".join(diff_summary) if diff_summary else "No textual changes detected."

#     # ----- Version number -----
#     version_number = doc.revisions.count() + 1

#     # ----- Save DOCX -----
#     folder = os.path.join(settings.MEDIA_ROOT, "qms_docs", "versions")
#     os.makedirs(folder, exist_ok=True)

#     output_filename = f"{doc.id}_v{version_number}.docx"
#     output_path = os.path.join(folder, output_filename)

#     try:
#         convert_html_to_docx(edited_html, output_path)
#     except Exception as e:
#         return JsonResponse({
#             "status": "error",
#             "message": f"DOCX generation failed: {str(e)}"
#         }, status=500)

#     # ----- Save Version -----
#     version = QMSDocumentVersion.objects.create(
#         document=doc,
#         version_number=version_number,
#         edited_by=request.user,
#         edited_html=edited_html,
#         edited_docx=f"qms_docs/versions/{output_filename}",
#     )

#     # ----- Audit -----
#     AuditLog.objects.create(
#         user=request.user,
#         role=request.user.role,
#         module="Documents",
#         action="UPDATE",
#         model_name="QMSDocument",
#         object_repr=doc.title,
#         new_data={"version": version_number},
#         ip_address=request.META.get("REMOTE_ADDR")
#     )

#     # ----- Save Revision -----
#     DocumentRevision.objects.create(
#         document=doc,
#         edited_by=request.user,
#         edited_html=edited_html,
#         change_summary=change_summary,
#         version_number=version_number
#     )

#     # Update title
#     if new_title and new_title != doc.title:
#         doc.title = new_title

#     doc.save()

#     return JsonResponse({
#         "status": "success",
#         "version": version_number
#     })
def generate_diff(old_html, new_html):
    from bs4 import BeautifulSoup
    import difflib

    def extract_text(html):
        return BeautifulSoup(html or "", "html.parser").get_text(separator=" ")

    old_text = extract_text(old_html)
    new_text = extract_text(new_html)

    old_words = old_text.split()
    new_words = new_text.split()

    changes = []

    sm = difflib.SequenceMatcher(None, old_words, new_words)

    for tag, i1, i2, j1, j2 in sm.get_opcodes():

        # 🔁 REPLACED (KEY FIX)
        if tag == "replace":
            old = " ".join(old_words[i1:i2]).strip()
            new = " ".join(new_words[j1:j2]).strip()

            if old and new and old != new:
                # avoid very large noisy output
                if len(old) < 100 and len(new) < 100:
                    changes.append(f"{old} → {new}")
                else:
                    changes.append("Content updated")

        # ➕ INSERT
        elif tag == "insert":
            new = " ".join(new_words[j1:j2]).strip()
            if new:
                if len(new) < 100:
                    changes.append(f"Added: {new}")
                else:
                    changes.append("Added content")

        # ❌ DELETE
        elif tag == "delete":
            old = " ".join(old_words[i1:i2]).strip()
            if old:
                if len(old) < 100:
                    changes.append(f"Removed: {old}")
                else:
                    changes.append("Removed content")

    # ✅ REMOVE DUPLICATES
    unique_changes = []
    for c in changes:
        if c not in unique_changes:
            unique_changes.append(c)

    return "\n".join(unique_changes) if unique_changes else "No changes detected"



from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.core.files import File
from django.utils import timezone
import json
import os


@login_required
def save_document(request, doc_id):

    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Invalid request method"},
            status=405
        )

    doc = get_object_or_404(QMSDocument, id=doc_id)

    # ================= READ DATA =================
    data = json.loads(request.body.decode("utf-8"))
    new_title = data.get("title", "").strip()

    # 🔥 DETECT FILE TYPE
    file_ext = os.path.splitext(doc.original_file.name)[1].lower()

    # ==========================================================
    # ✅ PDF FLOW (NO HTML REQUIRED)
    # ==========================================================
    if file_ext == ".pdf":

        signatures = data.get("pdf_signatures", [])
        new_title = data.get("title", "").strip()

        # ✅ Prevent accidental overwrite
        if signatures is None:
            signatures = []

        # ✅ Save signatures properly (MATCH FRONTEND)
        doc.pdf_signatures = signatures

        # ✅ Update title also
        if new_title and new_title != doc.title:
            doc.title = new_title

        doc.save()

        return JsonResponse({
            "status": "success",
            "message": "PDF signatures saved"
        })

    # ==========================================================
    # ✅ DOCX / HTML FLOW (OLD LOGIC CONTINUES)
    # ==========================================================
    edited_html = data.get("content", "").strip()

    if not edited_html:
        return JsonResponse({
            "status": "error",
            "message": "Content cannot be empty"
        }, status=400)
    # ==========================================================
    # 🔥 DRAFT MODE (ONLY REVISION, NO VERSION)
    # ==========================================================
    if doc.status != "completed":

        latest_revision = doc.revisions.order_by("-edited_at").first()

        if latest_revision and latest_revision.edited_html:
            old_html = latest_revision.edited_html
        else:
            latest_version = doc.versions.order_by("-version_number").first()
            old_html = latest_version.edited_html if latest_version else ""

        change_summary = generate_diff(old_html, edited_html)

        # ✅ Always create new revision
        DocumentRevision.objects.create(
            document=doc,
            edited_by=request.user,
            edited_html=edited_html,
            change_summary=change_summary
        )

        # ✅ Update title if changed
        if new_title and new_title != doc.title:
            doc.title = new_title
            doc.save()

        return JsonResponse({
            "status": "success",
            "message": "Draft saved (no version created)"
        })

    # ==========================================================
    # ✅ VERSION CREATION (COMPLETED DOCUMENT)
    # ==========================================================

    current_version = doc.versions.count()
    version_number = current_version + 1

    # ==========================================================
    # 🔥 ARCHIVE OLD VERSION (ONLY OLD, NOT NEW)
    # ==========================================================
    archive_path = os.path.join(settings.MEDIA_ROOT, "qms_docs", "archive")
    os.makedirs(archive_path, exist_ok=True)

    if current_version > 0 and doc.original_file:
        try:
            file_path = doc.original_file.path

            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    file_content = f.read()

                if file_content:
                    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

                    archive_filename = (
                        f"{doc.id}v{current_version}{timestamp}_"
                        f"{os.path.basename(file_path)}"
                    )

                    full_archive_path = os.path.join(archive_path, archive_filename)

                    with open(full_archive_path, "wb") as f:
                        f.write(file_content)

                    print("✅ ARCHIVED OLD VERSION:", archive_filename)

        except Exception as e:
            print("❌ ARCHIVE ERROR:", str(e))

    # ==========================================================
    # 🔥 GET OLD HTML (FOR DIFF)
    # ==========================================================
    latest_revision = doc.revisions.order_by("-edited_at").first()

    if latest_revision and latest_revision.edited_html:
        old_html = latest_revision.edited_html
    else:
        latest_version = doc.versions.order_by("-version_number").first()
        old_html = latest_version.edited_html if latest_version else ""

    change_summary = generate_diff(old_html, edited_html)

    # ==========================================================
    # 🔥 SAVE DOCX FILE
    # ==========================================================
    versions_path = os.path.join(settings.MEDIA_ROOT, "qms_docs", "versions")
    os.makedirs(versions_path, exist_ok=True)

    output_filename = f"{doc.id}_v{version_number}.docx"
    output_path = os.path.join(versions_path, output_filename)

    try:
        convert_html_to_docx(edited_html, output_path)
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"DOCX generation failed: {str(e)}"
        }, status=500)

    # ==========================================================
    # 🔥 SAVE VERSION ENTRY
    # ==========================================================
    version = QMSDocumentVersion.objects.create(
        document=doc,
        version_number=version_number,
        edited_by=request.user,
        edited_html=edited_html,
        edited_docx=f"qms_docs/versions/{output_filename}",
    )

    # ==========================================================
    # 🔥 UPDATE CURRENT FILE
    # ==========================================================
    try:
        with open(output_path, "rb") as f:
            doc.original_file.save(output_filename, File(f), save=False)
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"File save failed: {str(e)}"
        }, status=500)

    # ==========================================================
    # 🔥 UPDATE DOCUMENT
    # ==========================================================
    if new_title and new_title != doc.title:
        doc.title = new_title

    doc.save()

    # ==========================================================
    # 🔥 AUDIT LOG
    # ==========================================================
    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="Documents",
        action="UPDATE",
        model_name="QMSDocument",
        object_repr=doc.title,
        new_data={"version": version_number},
        ip_address=request.META.get("REMOTE_ADDR")
    )

    # ==========================================================
    # 🔥 SAVE REVISION (FOR HISTORY)
    # ==========================================================
    DocumentRevision.objects.create(
        document=doc,
        edited_by=request.user,
        edited_html=edited_html,
        change_summary=change_summary,
        version_number=version_number
    )

    return JsonResponse({
        "status": "success",
        "version": version_number
    })

@login_required
def archive_list(request):
    import os
    import re
    from django.conf import settings
    from django.urls import reverse

    archive_dir = os.path.join(settings.MEDIA_ROOT, "qms_docs", "archive")

    files = []

    if os.path.exists(archive_dir):
        for f in sorted(os.listdir(archive_dir), reverse=True):

            # Skip hidden + non-docx
            if f.startswith(".") or not f.endswith(".docx"):
                continue

            # ✅ Extract doc_id (only numbers at start)
            doc_match = re.match(r"(\d+)", f)
            doc_id = doc_match.group(1) if doc_match else "-"

            # ✅ Extract version (v5 → 5)
            version_match = re.search(r"_v(\d+)", f)
            version = version_match.group(1) if version_match else "-"

            # ✅ Extract timestamp (HHMMSS or similar)
            time_match = re.search(r"(\d{6})", f)
            timestamp = time_match.group(1) if time_match else "-"
            title = "-"
            if doc_id:
                doc = QMSDocument.objects.filter(id=doc_id).first()
                if doc:
                    title = doc.title
            files.append({
                "name": f,
                "title":title,
                "url": settings.MEDIA_URL + "qms_docs/archive/" + f,
                "view_url": reverse("view_archive_document", args=[f]),
                "doc_id": doc_id,
                "version": version,
                "timestamp": timestamp,
            })

    return render(request, "qms_app/sop/archive_list.html", {
        "files": files
    })
@login_required
def view_archive_document(request, filename):

    import re
    from .models import QMSDocumentVersion

    # ✅ Extract doc_id safely
    doc_match = re.match(r"(\d+)", filename)
    doc_id = doc_match.group(1) if doc_match else None

    # ✅ Extract version safely
    version_match = re.search(r"_v(\d+)", filename)
    version_no = version_match.group(1) if version_match else None

    if not doc_id or not version_no:
        return render(request, "qms_app/sop/error.html", {
            "message": "Invalid file format"
        })

    version = QMSDocumentVersion.objects.filter(
        document_id=int(doc_id),
        version_number=int(version_no)
    ).first()

    if not version:
        return render(request, "qms_app/sop/error.html", {
            "message": "Version not found"
        })

    html = version.edited_html

    page_divs = html.split('<div class="page-break"></div>')
    html_pages = [p.strip() for p in page_divs if p.strip()]

    return render(request, "qms_app/sop/view_archive.html", {
        "doc": {"title": filename},
        "html_pages": html_pages,
        "version": version_no
    })



@login_required
def document_workflow(request):
    access = UserPageAccess.objects.filter(user=request.user).first()
    return render(request, "qms_app/sop/document_workflow.html", {
        "process_docs": QMSDocument.objects.filter(status="process_owner"),
        "reviewed_docs": QMSDocument.objects.filter(status="reviewed"),
        "approved_docs": QMSDocument.objects.filter(status="approved"),
        "access": access,
    })
    
    
@login_required
def update_document_status(request, doc_id):

    doc = get_object_or_404(QMSDocument, id=doc_id)
    data = json.loads(request.body)

    new_status = data.get("status")
    user_id = data.get("user_id")
    message = data.get("message")

    doc.status = new_status

    # ✅ FIX: SET TARGET FOLDER
    if new_status == "approved":
        if not doc.target_folder:
            doc.target_folder = doc.folder

    # ================= ASSIGN USER =================
    if new_status in ["reviewed", "approved"]:

        if user_id:
            user = get_object_or_404(User, id=user_id)
            doc.assigned_to = user
        else:
            return JsonResponse({"error": "User required"}, status=400)

    elif new_status == "process_owner":
        doc.assigned_to = doc.created_by

    # ================= MESSAGE =================
    if message:
        doc.last_message = message
    else:
        if new_status == "reviewed":
            doc.last_message = "Sent for review"
        elif new_status == "approved":
            doc.last_message = "Sent for approval"
        elif new_status == "process_owner":
            doc.last_message = "Sent back for update"

    # ================= NOTIFICATION =================
    doc.is_read = False

    doc.save()
    DocumentWorkflowLog.objects.create(
        document=doc,
        stage=new_status,
        action="Moved",
        user=request.user,
        message=doc.last_message
    )
    return JsonResponse({"status": "ok"})


from django.contrib.auth import get_user_model
User = get_user_model()

@login_required
def send_back(request, doc_id):

    data = json.loads(request.body)
    msg = data.get("message")
    user_id = data.get("user_id")

    doc = get_object_or_404(QMSDocument, id=doc_id)
    user = get_object_or_404(User, id=user_id)

    doc.status = "process_owner"
    doc.assigned_to = user
    doc.last_message = msg or "Sent back for update"
    doc.is_read = False
    doc.save()

    return JsonResponse({"status": "ok"})


from django.utils.timezone import localtime

@login_required
def workflow_notifications(request):

    docs = QMSDocument.objects.filter(
        assigned_to=request.user,
        is_read=False
    )

    data = []

    for d in docs:
        data.append({
            "id": d.id,
            "title": d.title,
            "last_message": d.last_message,
            
            # ✅ ADD THESE
            "sender": d.created_by.email if d.created_by else "System",
            "time": localtime(d.created_at).strftime("%d %b %Y %I:%M %p")
        })

    return JsonResponse({"notifications": data})
    
    
# @login_required
# @require_POST
# def finalize_document(request, doc_id):

#     doc = get_object_or_404(QMSDocument, id=doc_id)

#     # ✅ AUTO FIX instead of error
#     if not doc.target_folder:
#         doc.target_folder = doc.folder

#     doc.folder = doc.target_folder
#     doc.status = "completed"
#     doc.assigned_to = None
#     doc.is_read = True
#     doc.save()

#     return JsonResponse({"success": True})
# @login_required
# @require_POST
# def finalize_document(request, doc_id):

#     doc = get_object_or_404(QMSDocument, id=doc_id)

#     if not doc.target_folder:
#         return JsonResponse({"error": "No target folder"}, status=400)

#     # ==========================================================
#     # 🔥 STEP 1: GET LATEST DRAFT (REVISION)
#     # ==========================================================
#     latest_revision = doc.revisions.order_by("-edited_at").first()

#     if latest_revision and latest_revision.edited_html:

#         edited_html = latest_revision.edited_html

#         # 🔢 Version number
#         version_number = doc.versions.count() + 1

#         # 📁 Save DOCX file
#         folder = os.path.join(settings.MEDIA_ROOT, "qms_docs", "versions")
#         os.makedirs(folder, exist_ok=True)

#         output_filename = f"{doc.id}_v{version_number}.docx"
#         output_path = os.path.join(folder, output_filename)

#         try:
#             convert_html_to_docx(edited_html, output_path)
#         except Exception as e:
#             print("DOCX ERROR:", str(e))

#         # 💾 CREATE VERSION
#         QMSDocumentVersion.objects.create(
#             document=doc,
#             version_number=version_number,
#             edited_by=request.user,
#             edited_html=edited_html,
#             edited_docx=f"qms_docs/versions/{output_filename}",
#         )

#     # ==========================================================
#     # 🔥 STEP 2: UPDATE DOCUMENT STATUS
#     # ==========================================================
#     doc.folder = doc.target_folder
#     doc.status = "completed"
#     doc.assigned_to = None
#     doc.is_read = True
#     doc.save()

#     return JsonResponse({"success": True})
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.core.files import File
from django.utils import timezone
import os


@login_required
@require_POST
def finalize_document(request, doc_id):

    import os
    from django.conf import settings
    from django.shortcuts import get_object_or_404
    from django.http import JsonResponse
    from django.utils import timezone
    from django.core.files import File

    doc = get_object_or_404(QMSDocument, id=doc_id)

    if not doc.target_folder:
        return JsonResponse({"error": "No target folder"}, status=400)

    # ==========================================================
    # 🔥 STEP 0: ARCHIVE OLD VERSION
    # ==========================================================
    archive_path = os.path.join(settings.MEDIA_ROOT, "qms_docs", "archive")
    os.makedirs(archive_path, exist_ok=True)

    current_version = doc.versions.count()

    if current_version > 0 and doc.original_file:
        try:
            file_path = doc.original_file.path

            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    file_content = f.read()

                if file_content:
                    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

                    archive_filename = (
                        f"{doc.id}v{current_version}{timestamp}_"
                        f"{os.path.basename(file_path)}"
                    )

                    full_archive_path = os.path.join(archive_path, archive_filename)

                    with open(full_archive_path, "wb") as f:
                        f.write(file_content)

                    print("✅ ARCHIVED FILE:", archive_filename)

                    # ✅ ALSO SAVE HTML (ONLY IF EXISTS)
                    latest_revision = doc.revisions.order_by("-edited_at").first()

                    if latest_revision and latest_revision.edited_html:
                        archive_html_filename = archive_filename.replace(".docx", ".html")
                        archive_html_path = os.path.join(archive_path, archive_html_filename)

                        with open(archive_html_path, "w", encoding="utf-8") as f:
                            f.write(latest_revision.edited_html)

                        print("✅ ARCHIVED HTML:", archive_html_filename)

        except Exception as e:
            print("❌ ARCHIVE ERROR:", str(e))

    # ==========================================================
    # 🔥 STEP 1: DETERMINE DOCUMENT TYPE
    # ==========================================================
    latest_revision = doc.revisions.order_by("-edited_at").first()

    has_html = latest_revision and latest_revision.edited_html
    has_pdf = doc.pdf_signatures and len(doc.pdf_signatures) > 0

    if not has_html and not has_pdf:
        return JsonResponse({"error": "No draft content found"}, status=400)

    edited_html = latest_revision.edited_html if latest_revision else ""
    version_number = current_version + 1

    versions_path = os.path.join(settings.MEDIA_ROOT, "qms_docs", "versions")
    os.makedirs(versions_path, exist_ok=True)

    # ==========================================================
    # 🔥 STEP 2: HANDLE DOCX vs PDF
    # ==========================================================
    if has_html:
        # ✅ DOCX FLOW
        output_filename = f"{doc.id}_v{version_number}.docx"
        output_path = os.path.join(versions_path, output_filename)

        try:
            convert_html_to_docx(edited_html, output_path)
        except Exception as e:
            return JsonResponse({
                "error": f"DOCX generation failed: {str(e)}"
            }, status=500)

    else:
        # ✅ PDF FLOW (NO CONVERSION)
        output_filename = os.path.basename(doc.original_file.name)
        output_path = doc.original_file.path

    # ==========================================================
    # 🔥 STEP 3: SAVE VERSION ENTRY
    # ==========================================================
    version = QMSDocumentVersion.objects.create(
        document=doc,
        version_number=version_number,
        edited_by=request.user,
        edited_html=edited_html,
        edited_docx=(
            f"qms_docs/versions/{output_filename}"
            if has_html else doc.original_file.name
        )
    )

    # ==========================================================
    # 🔥 STEP 4: UPDATE CURRENT FILE (ONLY FOR DOCX)
    # ==========================================================
    if has_html:
        try:
            with open(output_path, "rb") as f:
                doc.original_file.save(output_filename, File(f), save=False)
        except Exception as e:
            return JsonResponse({
                "error": f"File save failed: {str(e)}"
            }, status=500)

    # ==========================================================
    # 🔥 STEP 5: UPDATE DOCUMENT STATUS
    # ==========================================================
    doc.folder = doc.target_folder
    doc.status = "completed"
    doc.assigned_to = None
    doc.is_read = True
    doc.save()

    return JsonResponse({
        "success": True,
        "version": version_number
    })

@login_required
def workflow_data(request):

    def serialize(doc):
        return {
            "id": doc.id,
            "title": doc.title,
            "certificates": list(set(
                [c.name for c in doc.certificate.all()] +
                [cat.certificate.name for cat in doc.certificate_category.all() if cat.certificate]
            )),
            "categories": [c.name for c in doc.certificate_category.all()],

            "clause": doc.clause,

            "created_by": doc.created_by.email if doc.created_by else "Unknown",

            "folder": (
                doc.target_folder.name if doc.target_folder
                else doc.folder.name if doc.folder else "No Folder"
            ),

            "last_message": doc.last_message,
        }

    return JsonResponse({
        "process": [serialize(d) for d in QMSDocument.objects.filter(status="process_owner")],
        "reviewed": [serialize(d) for d in QMSDocument.objects.filter(status="reviewed")],
        "approved": [serialize(d) for d in QMSDocument.objects.filter(status="approved")],
    })
    
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse

User = get_user_model()

@login_required
def users_list(request):

    users = User.objects.all()

    data = []
    for u in users:
        data.append({
            "id": u.id,
            "name": u.email,
            "role": getattr(u, "role", "user")  # 🔥 SAFE fallback
        })

    return JsonResponse({"users": data})


@login_required
def mark_as_read(request, doc_id):
    doc = get_object_or_404(QMSDocument, id=doc_id)
    doc.is_read = True
    doc.save()
    return JsonResponse({"status": "ok"})
# ========================================================================================
# ================================ Delete Document =======================================
# ========================================================================================
@login_required
def delete_document(request, doc_id):
    doc = get_object_or_404(QMSDocument, id=doc_id)

    if request.method == "POST":
        doc.delete()
        messages.success(request, "Document deleted successfully.")
        return redirect("document_home")

    return redirect("document_home")




# ===========================================================================================================
# ============================================= Document Signature ==========================================
# ===========================================================================================================

def generate_sign_code():
    return "QMS-" + "-".join(
        ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        for _ in range(3)
    )


# ---------------- verify signature --------------------- 
from qms_app.models import UserPageAccess

@login_required
def verify_signature_password(request):

    if request.method != "POST":
        return JsonResponse({"status": "invalid"}, status=400)

    try:
        data = json.loads(request.body.decode())
    except:
        return JsonResponse({"status": "fail", "message": "Invalid data"}, status=400)

    password = data.get("password")
    sign_type = data.get("type")   # prepared / reviewed / approved / general

    # 🔐 Authenticate user
    user = authenticate(request, username=request.user.email, password=password)

    if not user:
        return JsonResponse({
            "status": "fail",
            "message": "Incorrect password"
        }, status=403)

    # ✅ AUTO CREATE ACCESS (FIX)
    access, created = UserPageAccess.objects.get_or_create(user=user)

    # 🔥 OPTIONAL: SUPERUSER FULL ACCESS
    if user.is_superuser:
        return JsonResponse({
            "status": "ok",
            "code": generate_sign_code(),
            "time": timezone.now().strftime("%d %b %Y %I:%M %p"),
            "user": request.user.get_display_name()
        })

    # 🔒 PERMISSION CHECK
    if sign_type == "prepared" and not access.can_sign_prepared:
        return JsonResponse({
            "status": "fail",
            "message": "You are not allowed to sign as Prepared By"
        }, status=403)

    if sign_type == "reviewed" and not access.can_sign_reviewed:
        return JsonResponse({
            "status": "fail",
            "message": "You are not allowed to sign as Reviewed By"
        }, status=403)

    if sign_type == "approved" and not access.can_sign_approved:
        return JsonResponse({
            "status": "fail",
            "message": "You are not allowed to sign as Approved By"
        }, status=403)

    # 🔥 GENERAL SIGN (TEXT DOCS) → allow if any permission
    if sign_type == "general":
        if not (access.can_sign_prepared or access.can_sign_reviewed or access.can_sign_approved):
            return JsonResponse({
                "status": "fail",
                "message": "You are not allowed to sign this document"
            }, status=403)

    # ✅ SUCCESS
    return JsonResponse({
        "status": "ok",
        "code": generate_sign_code(),
        "time": timezone.now().strftime("%d %b %Y %I:%M %p"),
        "user": request.user.get_display_name()
    })


# -------------------- generate code ----------------------
def generate_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))

# -------------------- sign document ----------------------
@login_required
def sign_document(request, doc_id):

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST only"})

    data = json.loads(request.body.decode("utf-8"))
    password = data.get("password")

    # VERIFY PASSWORD
    if not check_password(password, request.user.password):
        return JsonResponse({
            "status": "error",
            "message": "Invalid password"
        })

    code = generate_code()

    return JsonResponse({
        "status": "ok",
        "user": request.user.get_display_name(),
        "time": timezone.now().strftime("%d-%m-%Y %I:%M %p"),
        "code": code
    })






# ========================================================================================================================================
# #######################################################       Form Creation     ########################################################
# ========================================================================================================================================
# @require_page_permission("can_forms")
# @login_required
# def form_list(request):

#     folder_id = request.GET.get("folder")
#     selected_folder = None

#     # =====================================
#     #  SUPERUSER = FULL ACCESS
#     # =====================================
#     if request.user.is_superuser:

#         # Only root folders for sidebar
#         root_folders = (
#             FormFolder.objects
#             .filter(parent__isnull=True)
#             .prefetch_related("children")
#         )

#         forms = Form.objects.all()

#         if folder_id:
#             selected_folder = FormFolder.objects.filter(id=folder_id).first()
#             if selected_folder:
#                 forms = forms.filter(folder=selected_folder)

#         active_batches = (
#             FormBatch.objects
#             .select_related("current_stage")
#             .order_by("-created_at")
#         )

#         return render(request, "qms_app/form/form_list.html", {
#             "forms": forms,
#             "folders": root_folders,
#             "selected_folder": selected_folder,
#             "active_batches": active_batches
#         })


#     # =====================================
#     #  NORMAL USER ACCESS CONTROL
#     # =====================================

#     page_access, created = UserPageAccess.objects.get_or_create(
#         user=request.user
#     )

#     #  Page-level restriction
#     if not page_access.can_forms:
#         return redirect("denied")

#     # Only folders this user is allowed to see
#     allowed_folders = page_access.allowed_folders.all()

#     # Sidebar should show only ROOT allowed folders
#     root_folders = (
#         allowed_folders
#         .filter(parent__isnull=True)
#         .prefetch_related("children")
#     )

#     # -------------------------------------

#     if folder_id:
#         selected_folder = allowed_folders.filter(id=folder_id).first()

#         if not selected_folder:
#             return redirect("denied")

#         forms = Form.objects.filter(folder=selected_folder)
#     else:
#         # Show forms inside ANY allowed folder
#         forms = Form.objects.filter(folder__in=allowed_folders)

#     active_batches = (
#         FormBatch.objects
#         .select_related("current_stage")
#         .order_by("-created_at")
#     )

#     return render(request, "qms_app/form/form_list.html", {
#         "forms": forms,
#         "folders": root_folders,
#         "selected_folder": selected_folder,
#         "active_batches": active_batches
#     })


@require_page_permission("can_forms")
@login_required
def form_list(request):

    folder_id = request.GET.get("folder")
    selected_folder = None
    folder_path = []

    # =====================================
    # SUPERUSER = FULL ACCESS
    # =====================================
    if request.user.is_superuser:

        root_folders = (
            FormFolder.objects
            .filter(parent__isnull=True)
            .prefetch_related("children")
        )

        forms = Form.objects.all()

        if folder_id:
            selected_folder = FormFolder.objects.filter(id=folder_id).first()

            if selected_folder:

                # Build breadcrumb path
                current = selected_folder
                while current:
                    folder_path.insert(0, current)
                    current = current.parent

                # Show only forms inside this folder
                forms = forms.filter(folder=selected_folder)

        active_batches = (
            FormBatch.objects
            .select_related("current_stage")
            .order_by("-created_at")
        )

        return render(request, "qms_app/form/form_list.html", {
            "forms": forms,
            "folders": root_folders,
            "selected_folder": selected_folder,
            "folder_path": folder_path,
            "active_batches": active_batches
        })


    # =====================================
    # NORMAL USER ACCESS CONTROL
    # =====================================

    page_access, created = UserPageAccess.objects.get_or_create(
        user=request.user
    )

    if not page_access.can_forms:
        return redirect("denied")

    selected_folders = page_access.allowed_folders.all()

    def get_all_descendants(folder):
        children = list(folder.children.all())
        for child in folder.children.all():
            children += get_all_descendants(child)
        return children

    all_allowed = []

    for folder in selected_folders:
        all_allowed.append(folder)
        all_allowed += get_all_descendants(folder)

    allowed_folders = FormFolder.objects.filter(
        id__in=[f.id for f in all_allowed]
    )

    root_folders = FormFolder.objects.filter(
        id__in=[f.id for f in all_allowed],
        parent__isnull=True
    ).prefetch_related("children")

    if folder_id:

        selected_folder = allowed_folders.filter(id=folder_id).first()

        if not selected_folder:
            return redirect("denied")

        current = selected_folder
        while current:
            folder_path.insert(0, current)
            current = current.parent

        forms = Form.objects.filter(folder=selected_folder)

    else:
        forms = Form.objects.filter(folder__in=allowed_folders)

    active_batches = (
        FormBatch.objects
        .select_related("current_stage")
        .order_by("-created_at")
    )

    return render(request, "qms_app/form/form_list.html", {
        "forms": forms,
        "folders": root_folders,
        "selected_folder": selected_folder,
        "folder_path": folder_path,
        "active_batches": active_batches
    })
# ==================================================================================================
# ====================================== batch Tracker =============================================
# ==================================================================================================

@login_required
def tracker_search(request):

    company = request.GET.get("company", "").strip()
    batch = request.GET.get("batch", "").strip()

    queryset = FormBatch.objects.filter(is_active=True)

    if company:
        queryset = queryset.filter(company_name__icontains=company)

    if batch:
        queryset = queryset.filter(batch_id__icontains=batch)

    queryset = queryset.select_related("form", "current_stage", "created_by")

    results = []

    for obj in queryset:
        results.append({
            "company_name": obj.company_name,
            "batch_id": obj.batch_id,
            "form_name": obj.form.name,
            "stage_name": obj.current_stage.name,
            "created_by": obj.created_by.email,
            "created_at": obj.created_at.strftime("%d %b %Y, %H:%M"),
        })

    return JsonResponse({"results": results})



# --- Create Form ---
class FormCreateForm(forms.ModelForm):
    class Meta:
        model = Form
        fields = ['name', 'description']


# ==================================================================================================
# ========================================== Form Creation =========================================
# ==================================================================================================

@login_required
def create_form(request):

    folder_id = request.GET.get("folder")
    selected_folder = None

    if folder_id:
        selected_folder = FormFolder.objects.filter(id=folder_id).first()

    if request.method == "POST":
        form = FormCreateForm(request.POST)

        if form.is_valid():
            new_form = form.save(commit=False)
            new_form.created_by = request.user
            new_form.folder = selected_folder
            new_form.save()

            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="Form",
                action="CREATE",
                model_name="Form",
                object_repr=new_form.name,
                new_data={"description": new_form.description},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect("add_stage", form_id=new_form.id)

    else:
        form = FormCreateForm()

    return render(
        request,
        "qms_app/form/create_form.html",
        {
            "form": form,
            "selected_folder": selected_folder
        }
    )




# ------------------------- create form folder ----------------------------
import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import FormFolder   # ✅ correct model

@login_required
def create_form_folder(request):
    if request.method == "POST":
        data = json.loads(request.body)

        name = data.get('name')
        parent_id = data.get('parent')

        if parent_id in ["", None, "null"]:
            parent = None
        else:
            parent = FormFolder.objects.get(id=int(parent_id))  # ✅ correct model

        if name:
            FormFolder.objects.create(
                name=name,
                created_by=request.user,
                parent=parent
            )

        return JsonResponse({"success": True})

    return JsonResponse({"success": False})

# ------------------------- delete form folder -----------------------------
@login_required
def delete_form_folder(request, folder_id):

    folder = FormFolder.objects.filter(id=folder_id).first()

    if folder:
        folder.delete()

    return redirect("form_list")

# --------------------------- form edit -----------------------------------
@login_required
def edit_form(request, form_id):
    form_obj = get_object_or_404(Form, id=form_id)

    if request.method == 'POST':
        form = FormCreateForm(request.POST, instance=form_obj)
        if form.is_valid():
            form.save()
            return redirect('form_list')
    else:
        form = FormCreateForm(instance=form_obj)

    return render(request, 'qms_app/form/edit_form.html', {
        'form_obj': form_obj,
        'form': form
    })

# --------------------------- form delete -----------------------------------
def delete_form(request, form_id):
    form = get_object_or_404(Form, id=form_id)
    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="Form",
        action="DELETE",
        model_name="Form",
        object_repr=form.name,
        old_data={"id": form.id},
        ip_address=request.META.get("REMOTE_ADDR")
    )

    form.delete()
    return redirect('form_list')


# =============================================================================================
# ===================================== Add Stage =============================================
# =============================================================================================

@login_required
def add_stage(request, form_id):
    form_obj = get_object_or_404(Form, id=form_id)

    #  Only admin can add stages
    if request.user.role != "admin":
        return HttpResponseForbidden("You are not allowed to add stages.")

    if request.method == 'POST':
        stage_name = request.POST.get('name')
        allowed_roles = request.POST.getlist('allowed_roles')  #  REQUIRED

        if stage_name and allowed_roles:
            Stage.objects.create(
                form=form_obj,
                name=stage_name,
                order=form_obj.stages.count() + 1,  # cleaner ordering
                allowed_roles=allowed_roles
            )
            
            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="Stage",
                action="CREATE",
                model_name="Stage",
                object_repr=stage_name,
                new_data={"allowed_roles": allowed_roles},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect('add_stage', form_id=form_obj.id)

    stages = form_obj.stages.all()

    return render(request, 'qms_app/form/add_stage.html', {
        'form_obj': form_obj,
        'stages': stages
    })




# ==========================================================================================
# ================================== edit stage =========================================
# =======================================================================================

@login_required
def edit_stage(request, stage_id):
    stage = get_object_or_404(Stage, id=stage_id)

    #  Admin only
    if request.user.role != "admin":
        return HttpResponseForbidden("You are not allowed to edit stages.")

    role_choices = CustomUser.ROLE_CHOICES  #  SAFE

    if request.method == "POST":
        stage_name = request.POST.get("name")
        allowed_roles = request.POST.getlist("allowed_roles")

        if stage_name and allowed_roles:
            stage.name = stage_name
            stage.allowed_roles = allowed_roles
            stage.save()
            return redirect("add_stage", form_id=stage.form.id)

    return render(request, "qms_app/form/edit_stage.html", {
        "stage": stage,
        "role_choices": role_choices
    })




# ====================================================================================
# ================================== Add Field =======================================
# ====================================================================================
# @login_required
# def add_field(request, stage_id):
#     stage = get_object_or_404(Stage, id=stage_id)

#     main_fields = stage.fields.filter(
#         sub_stage__isnull=True
#     ).order_by("order")

#     sub_stages = stage.sub_stages.prefetch_related(
#         "fields"
#     ).order_by("order")

#     if request.method == "POST":

#         label = request.POST.get("label", "").strip()
#         field_type = request.POST.get("field_type", "").strip()
#         options_raw = request.POST.get("options", "").strip()
#         table_columns_raw = request.POST.get("table_columns", "")
#         table_rows_raw = request.POST.get("table_rows", "")
#         table_row_header = request.POST.get("table_row_header", "").strip()  #  NEW
#         sub_stage_id = request.POST.get("sub_stage")
#         new_sub_stage = request.POST.get("new_sub_stage", "").strip()

#         VALID_FIELD_TYPES = {
#             "text", "textarea", "number",
#             "date", "select", "checkbox", "table", "image_upload", "image_capture"
#         }

#         if field_type not in VALID_FIELD_TYPES:
#             return redirect("add_field", stage_id=stage.id)

#         # ---------------- Resolve SubStage ----------------
#         sub_stage = None

#         if new_sub_stage:
#             sub_stage = SubStage.objects.create(
#                 stage=stage,
#                 name=new_sub_stage,
#                 order=(
#                     stage.sub_stages.aggregate(Max("order"))["order__max"] or 0
#                 ) + 1
#             )
#         elif sub_stage_id:
#             sub_stage = get_object_or_404(
#                 SubStage,
#                 id=sub_stage_id,
#                 stage=stage
#             )

#         # ---------------- Select / Checkbox ----------------
#         options = ""
#         if field_type in {"select", "checkbox"}:
#             options_list = [
#                 x.strip() for x in options_raw.split(",") if x.strip()
#             ]
#             options = ",".join(options_list)

#         # ---------------- TABLE HANDLING ----------------
#         table_columns = None
#         table_rows = None

#         if field_type == "table":
#             try:
#                 table_columns = json.loads(table_columns_raw) if table_columns_raw else []
#                 table_rows = json.loads(table_rows_raw) if table_rows_raw else []
#             except json.JSONDecodeError:
#                 return redirect("add_field", stage_id=stage.id)

#             if not isinstance(table_columns, list) or not isinstance(table_rows, list):
#                 return redirect("add_field", stage_id=stage.id)

#             cleaned_columns = []
#             for col in table_columns:
#                 name = (col.get("name") or "").strip()
#                 if not name:
#                     continue

#                 cleaned_columns.append({
#                     "name": name,
#                     "type": col.get("type", "text"),
#                     "options": col.get("options", []),
#                     "editable": True,
#                 })

#             cleaned_rows = []
#             for row in table_rows:
#                 row_name = (row.get("name") or "").strip()
#                 cells = row.get("cells", [])

#                 if not row_name:
#                     continue

#                 if len(cells) < len(cleaned_columns):
#                     cells += [""] * (len(cleaned_columns) - len(cells))
#                 elif len(cells) > len(cleaned_columns):
#                     cells = cells[:len(cleaned_columns)]

#                 cleaned_rows.append({
#                     "name": row_name,
#                     "cells": cells
#                 })

#             table_columns = cleaned_columns
#             table_rows = cleaned_rows

#         # ---------------- SAVE FIELD ----------------
#         if label:
#             next_order = (
#                 stage.fields.aggregate(Max("order"))["order__max"] or 0
#             ) + 1

#             FormField.objects.create(
#                 stage=stage,
#                 sub_stage=sub_stage,
#                 label=label,
#                 field_type=field_type,
#                 options=options,
#                 table_columns=table_columns,
#                 table_rows=table_rows,
#                 table_row_header=table_row_header,  #  SAVED
#                 order=next_order
#             )
#             AuditLog.objects.create(
#                 user=request.user,
#                 role=request.user.role,
#                 module="Field",
#                 action="CREATE",
#                 model_name="FormField",
#                 object_repr=label,
#                 new_data={
#                     "stage": stage.name,
#                     "sub_stage": sub_stage.name if sub_stage else None,
#                     "field_type": field_type,
#                     "options": options,
#                     "table_columns": table_columns,
#                     "table_rows": table_rows
#                 },
#                 ip_address=request.META.get("REMOTE_ADDR")
#             )

#         return redirect("add_field", stage_id=stage.id)

#     return render(
#         request,
#         "qms_app/form/add_field.html",
#         {
#             "stage": stage,
#             "main_fields": main_fields,
#             "sub_stages": sub_stages,
#         }
#     )
@login_required
def add_field(request, stage_id):
    stage = get_object_or_404(Stage, id=stage_id)

    main_fields = stage.fields.filter(
        sub_stage__isnull=True
    ).order_by("order")

    sub_stages = stage.sub_stages.prefetch_related(
        "fields"
    ).order_by("order")

    if request.method == "POST":

        label = request.POST.get("label", "").strip()
        field_type = request.POST.get("field_type", "").strip()
        options_raw = request.POST.get("options", "").strip()

        table_columns_raw = request.POST.get("table_columns", "")
        table_rows_raw = request.POST.get("table_rows", "")
        table_row_header = request.POST.get("table_row_header", "").strip()

        sub_stage_id = request.POST.get("sub_stage")
        new_sub_stage = request.POST.get("new_sub_stage", "").strip()

        VALID_FIELD_TYPES = {
            "text", "textarea", "number",
            "date", "select", "checkbox",
            "table", "image_upload", "image_capture"
        }

        if field_type not in VALID_FIELD_TYPES:
            return redirect("add_field", stage_id=stage.id)

        # ================= SUB STAGE =================
        sub_stage = None

        if new_sub_stage:
            sub_stage = SubStage.objects.create(
                stage=stage,
                name=new_sub_stage,
                order=(
                    stage.sub_stages.aggregate(Max("order"))["order__max"] or 0
                ) + 1
            )

        elif sub_stage_id:
            sub_stage = get_object_or_404(
                SubStage,
                id=sub_stage_id,
                stage=stage
            )

        # ================= OPTIONS =================
        options = ""
        if field_type in {"select", "checkbox"}:
            options_list = [
                x.strip() for x in options_raw.split(",") if x.strip()
            ]
            options = ",".join(options_list)

        # ================= TABLE =================
        table_columns = None
        table_rows = None

        if field_type == "table":

            try:
                table_columns = json.loads(table_columns_raw) if table_columns_raw else []
                table_rows = json.loads(table_rows_raw) if table_rows_raw else []
            except json.JSONDecodeError:
                return redirect("add_field", stage_id=stage.id)

            if not isinstance(table_columns, list) or not isinstance(table_rows, list):
                return redirect("add_field", stage_id=stage.id)

            # ================= CLEAN COLUMNS =================
            cleaned_columns = []

            for col in table_columns:

                name = (col.get("name") or "").strip()
                if not name:
                    continue

                col_type = col.get("type", "text")
                formula_value = (col.get("formula") or "").strip()
                if formula_value:
                    col_type = "formula"
                VALID_COL_TYPES = ["text", "number", "percentage", "formula"]

                if col_type not in VALID_COL_TYPES:
                    col_type = "text"

                cleaned_columns.append({
                    "name": name,
                    "type": col_type,
                    "options": col.get("options", []),
                    "editable": True,

                    # 🔥 KEEP THIS
                    "is_total": bool(col.get("is_total", False)),
                    "formula": (col.get("formula") or "").strip(),
                })

            # ================= CLEAN ROWS (🔥 FIXED) =================
            cleaned_rows = []

            for row in table_rows:

                row_name = (row.get("name") or "").strip()
                raw_cells = row.get("cells", [])

                if not row_name:
                    continue

                cleaned_cells = []

                for i, col in enumerate(cleaned_columns):

                    col_type = col.get("type", "text")

                    cell = raw_cells[i] if i < len(raw_cells) else {}

                    # 🔥 IMPORTANT FIX
                    formula_val = ""

                    if isinstance(cell, dict):
                        formula_val = cell.get("formula")

                    if not formula_val:
                        formula_val = col.get("formula")

                    if isinstance(cell, dict):
                        cleaned_cells.append({
                            "type": cell.get("type", col_type),
                            "value": cell.get("value", ""),
                            "options": cell.get("options", []),

                            # 🔥 PRESERVE FORMULA
                            "formula": formula_val or "",
                        })
                    else:
                        cleaned_cells.append({
                            "type": col_type,
                            "value": "",
                            "options": [],

                            # 🔥 PRESERVE FORMULA FROM COLUMN
                            "formula": formula_val or "",
                        })

                cleaned_rows.append({
                    "name": row_name,
                    "cells": cleaned_cells
                })

            table_columns = cleaned_columns
            table_rows = cleaned_rows

        # ================= SAVE =================
        if label:

            next_order = (
                stage.fields.aggregate(Max("order"))["order__max"] or 0
            ) + 1

            FormField.objects.create(
                stage=stage,
                sub_stage=sub_stage,
                label=label,
                field_type=field_type,
                options=options,
                table_columns=table_columns,
                table_rows=table_rows,
                table_row_header=table_row_header,
                order=next_order
            )

            # ================= AUDIT =================
            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="Field",
                action="CREATE",
                model_name="FormField",
                object_repr=label,
                new_data={
                    "stage": stage.name,
                    "sub_stage": sub_stage.name if sub_stage else None,
                    "field_type": field_type,
                    "options": options,
                    "table_columns": table_columns,
                    "table_rows": table_rows
                },
                ip_address=request.META.get("REMOTE_ADDR")
            )

        return redirect("add_field", stage_id=stage.id)

    return render(
        request,
        "qms_app/form/add_field.html",
        {
            "stage": stage,
            "main_fields": main_fields,
            "sub_stages": sub_stages,
        }
    )
# ==================================================================================================
# ========================================== Delete Field ==========================================
# ==================================================================================================
@login_required
def delete_field(request, field_id):
    field = get_object_or_404(FormField, id=field_id)
    stage_id = field.stage.id

    if request.method == "POST":
        field.delete()
        return redirect("add_field", stage_id=stage_id)

    return render(request, "qms_app/form/confirm_delete.html", {
        "field": field,
        "stage_id": stage_id
    })



# ---------------------------- Edit Field --------------------------------
# @login_required
# def edit_field(request, field_id):
#     field = get_object_or_404(FormField, id=field_id)
#     stage = field.stage
#     error = None

#     VALID_FIELD_TYPES = [
#         "text", "textarea", "number",
#         "date", "select", "checkbox", "table"
#     ]

#     if request.method == "POST":

#         label = request.POST.get("label", "").strip()
#         field_type = request.POST.get("field_type", "").strip()
#         options_raw = request.POST.get("options", "").strip()

#         table_columns_raw = request.POST.get("table_columns", "")
#         table_rows_raw = request.POST.get("table_rows", "")

#         if field_type not in VALID_FIELD_TYPES:
#             error = "Invalid field type."

#         elif not label:
#             error = "Label is required."

#         else:

#             field.label = label
#             field.field_type = field_type

#             # --------------------------
#             # SELECT / CHECKBOX
#             # --------------------------
#             if field_type in ["select", "checkbox"]:
#                 options_list = [
#                     x.strip() for x in options_raw.split(",")
#                     if x.strip()
#                 ]
#                 field.options = ",".join(options_list)
#                 field.table_columns = None
#                 field.table_rows = None

#             # --------------------------
#             # TABLE
#             # --------------------------
#             elif field_type == "table":

#                 try:
#                     table_columns = json.loads(table_columns_raw) if table_columns_raw else []
#                     table_rows = json.loads(table_rows_raw) if table_rows_raw else []
#                 except json.JSONDecodeError:
#                     error = "Invalid table format."
#                     return render(request, "qms_app/form/edit_field.html", {
#                         "field": field,
#                         "stage": stage,
#                         "error": error
#                     })

#                 cleaned_columns = []
#                 for col in table_columns:
#                     name = (col.get("name") or "").strip()
#                     if not name:
#                         continue

#                     cleaned_columns.append({
#                         "name": name,
#                         "type": col.get("type", "text"),
#                         "options": col.get("options", []),
#                         "editable": col.get("editable", True)
#                     })

#                 cleaned_rows = []
#                 for row in table_rows:
#                     row_name = (row.get("name") or "").strip()
#                     cells = row.get("cells", [])

#                     if not row_name:
#                         continue

#                     if len(cells) < len(cleaned_columns):
#                         cells += [""] * (len(cleaned_columns) - len(cells))
#                     elif len(cells) > len(cleaned_columns):
#                         cells = cells[:len(cleaned_columns)]

#                     cleaned_rows.append({
#                         "name": row_name,
#                         "cells": cells
#                     })

#                 field.table_columns = cleaned_columns
#                 field.table_rows = cleaned_rows
#                 field.options = ""

#             # --------------------------
#             # NORMAL FIELDS
#             # --------------------------
#             else:
#                 field.options = ""
#                 field.table_columns = None
#                 field.table_rows = None

#             field.save()
#             return redirect("add_field", stage_id=stage.id)

#     return render(request, "qms_app/form/edit_field.html", {
#         "field": field,
#         "stage": stage,
#         "error": error
#     })

@login_required
def edit_field(request, field_id):

    field = get_object_or_404(FormField, id=field_id)
    stage = field.stage
    error = None

    VALID_FIELD_TYPES = [
        "text", "textarea", "number",
        "date", "select", "checkbox", "table"
    ]

    if request.method == "POST":

        label = request.POST.get("label", "").strip()
        field_type = request.POST.get("field_type", "").strip()
        options_raw = request.POST.get("options", "").strip()

        table_columns_raw = request.POST.get("table_columns", "")
        table_rows_raw = request.POST.get("table_rows", "")

        if field_type not in VALID_FIELD_TYPES:
            error = "Invalid field type."

        elif not label:
            error = "Label is required."

        else:

            field.label = label
            field.field_type = field_type

            # ================= SELECT / CHECKBOX =================
            if field_type in ["select", "checkbox"]:

                options_list = [
                    x.strip() for x in options_raw.split(",") if x.strip()
                ]

                field.options = ",".join(options_list)
                field.table_columns = None
                field.table_rows = None

            # ================= TABLE =================
            elif field_type == "table":

                try:
                    table_columns = json.loads(table_columns_raw) if table_columns_raw else []
                    table_rows = json.loads(table_rows_raw) if table_rows_raw else []
                except json.JSONDecodeError:
                    error = "Invalid table format."
                    return render(request, "qms_app/form/edit_field.html", {
                        "field": field,
                        "stage": stage,
                        "error": error
                    })

                # ================= CLEAN COLUMNS =================
                cleaned_columns = []

                VALID_COL_TYPES = [
                    "text", "number", "percentage",
                    "formula", "checkbox", "select", "date"
                ]

                for col in table_columns:

                    name = (col.get("name") or "").strip()
                    if not name:
                        continue

                    col_type = col.get("type", "text").lower()

                    if col_type not in VALID_COL_TYPES:
                        col_type = "text"

                    cleaned_columns.append({
                        "name": name,
                        "type": col_type,
                        "options": col.get("options", []),
                        "editable": col.get("editable", True),
                        "is_total": bool(col.get("is_total", False)),
                        "formula": (col.get("formula") or "").strip(),
                    })

                # ================= CLEAN ROWS =================
                cleaned_rows = []

                for row in table_rows:

                    row_name = (row.get("name") or "").strip()
                    raw_cells = row.get("cells", [])

                    if not row_name:
                        continue

                    cleaned_cells = []

                    for i, col in enumerate(cleaned_columns):

                        col_type = col.get("type", "text")

                        # existing cell
                        cell = raw_cells[i] if i < len(raw_cells) else {}

                        if isinstance(cell, dict):
                            cleaned_cells.append({
                                "type": cell.get("type", col_type),
                                "value": cell.get("value", ""),
                                "options": cell.get("options", []),
                                "formula": cell.get("formula", col.get("formula", ""))
                            })
                        else:
                            # fallback
                            cleaned_cells.append({
                                "type": col_type,
                                "value": "",
                                "options": [],
                                "formula": col.get("formula", "")
                            })

                    cleaned_rows.append({
                        "name": row_name,
                        "cells": cleaned_cells
                    })

                field.table_columns = cleaned_columns
                field.table_rows = cleaned_rows
                field.options = ""

            # ================= NORMAL FIELD =================
            else:
                field.options = ""
                field.table_columns = None
                field.table_rows = None

            field.save()

            return redirect("add_field", stage_id=stage.id)

    return render(request, "qms_app/form/edit_field.html", {
        "field": field,
        "stage": stage,
        "error": error
    })


# ===================================================================================================
# ======================================= From Generation Logic =====================================
# ===================================================================================================
import re

def normalize_key(name):
    return re.sub(r'[^a-z0-9]', '', name.lower())

from django import forms
import json

def generate_form(stage, data=None, initial=None, files=None):

    fields = (
        stage.fields
        .select_related("sub_stage")
        .all()
        .order_by("sub_stage__order", "order")
    )

    form_fields = {}

    for field in fields:
        field_name = f"field_{field.id}"

        common_kwargs = {
            "label": field.label,
            "required": field.required,
        }

        # ---------------- NORMAL FIELDS ---------------- #

        if field.field_type == "text":
            form_field = forms.CharField(**common_kwargs)

        elif field.field_type == "textarea":
            form_field = forms.CharField(
                widget=forms.Textarea(attrs={"rows": 3}),
                **common_kwargs
            )

        elif field.field_type == "number":
            form_field = forms.IntegerField(**common_kwargs)

        elif field.field_type == "date":
            form_field = forms.DateField(
                widget=forms.DateInput(attrs={"type": "date"}),
                **common_kwargs
            )

        elif field.field_type == "select":

            raw = field.options

            if isinstance(raw, list):
                options_list = raw
            elif isinstance(raw, str):
                options_list = [x.strip() for x in raw.split(",") if x.strip()]
            else:
                options_list = []

            choices = [(opt, opt) for opt in options_list]

            if not choices:
                choices = [("N/A", "N/A")]

            form_field = forms.ChoiceField(
                choices=choices,
                **common_kwargs
            )

            form_field.widget.attrs["is_select"] = True

        elif field.field_type == "checkbox":

            raw = field.options

            if isinstance(raw, list):
                choices = [(x, x) for x in raw]
            elif isinstance(raw, str):
                choices = [(x.strip(), x.strip()) for x in raw.split(",") if x.strip()]
            else:
                choices = []

            if choices:
                form_field = forms.MultipleChoiceField(
                    choices=choices,
                    widget=forms.CheckboxSelectMultiple,
                    required=field.required,
                    label=field.label
                )
            else:
                form_field = forms.BooleanField(
                    required=field.required,
                    label=field.label
                )

            form_field.widget.attrs["is_checkbox"] = True

        elif field.field_type == "image_upload":

            form_field = forms.ImageField(
                required=field.required,
                label=field.label
            )

            form_field.widget.attrs.update({
                "accept": "image/*"
            })

        elif field.field_type == "image_capture":

            form_field = forms.ImageField(
                required=field.required,
                label=field.label
            )

            form_field.widget.attrs.update({
                "accept": "image/*",
                "capture": "environment"
            })

        # ================= TABLE FIELD (FULL UPGRADE) ================= #

        elif field.field_type == "table":

            form_field = forms.CharField(
                required=False,
                widget=forms.HiddenInput()
            )

            columns = field.table_columns or []
            rows = field.table_rows or []

            submitted_data = None

            if initial and field_name in initial:
                submitted_data = initial.get(field_name)

                if isinstance(submitted_data, str):
                    try:
                        submitted_data = json.loads(submitted_data)
                    except Exception:
                        submitted_data = None

            # ================= NORMALIZE COLUMNS =================
            normalized_columns = []

            for col in columns:

                if isinstance(col, dict):

                    normalized_columns.append({
                        "name": col.get("name", ""),
                        "type": col.get("type", "text"),

                        "options": col.get("options", []),
                        "is_total": str(col.get("is_total")).lower() == "true",
                        "formula": col.get("formula", ""),
                    })

                else:
                    normalized_columns.append({
                        "name": str(col),
                        "type": "text",
                        "is_total": False,
                        "formula": "",
                    })

            # ================= NORMALIZE ROWS =================
            normalized_rows = []

            for row_index, row in enumerate(rows):

                if isinstance(row, str):
                    row_name = row
                    original_cells = []

                elif isinstance(row, dict):
                    row_name = row.get("name", "")
                    original_cells = row.get("cells", [])

                else:
                    row_name = str(row)
                    original_cells = []

                # override with submitted data
                if submitted_data and row_index < len(submitted_data):
                    saved_row = submitted_data[row_index]
                    original_cells = saved_row.get("cells", [])

                cell_objects = []

                for col_index, col in enumerate(normalized_columns):

                    col_type = col.get("type", "text")

                    if (
                        isinstance(original_cells, list)
                        and col_index < len(original_cells)
                        and isinstance(original_cells[col_index], dict)
                    ):
                        cell_data = original_cells[col_index]

                        cell_type = cell_data.get("type") or col.get("type", "text")

                        cell_objects.append({
                            "type": cell_type,   # 🔥 ROW PRIORITY
                            "options": cell_data.get("options") or col.get("options", []),
                            "value": cell_data.get("value", ""),
                            "colname": normalize_key(col.get("name", "")),
                            "formula": cell_data.get("formula") or col.get("formula", "")
                        })

                    else:
                        cell_objects.append({
                            "type": col.get("type", "text"),
                            "options": col.get("options", []),
                            "value": "",
                            "colname": normalize_key(col.get("name", "")),
                            "formula": col.get("formula", "")
                        })

                normalized_rows.append({
                    "name": row_name,
                    "cells": cell_objects
                })
            has_total = any(col.get("is_total") for col in normalized_columns)
            # ================= SEND TO FRONTEND =================
            form_field.widget.attrs.update({
                "is_table": True,
                "columns": normalized_columns,
                "columns_json": json.dumps(normalized_columns),
                "rows": normalized_rows,
                "field_label": field.label,
                "row_header": getattr(field, "table_row_header", ""),
                "has_total": has_total,
                
            })

        # ---------------- DEFAULT ---------------- #
        else:
            form_field = forms.CharField(**common_kwargs)

        form_field.widget.attrs["substage"] = (
            str(field.sub_stage_id)
            if field.sub_stage_id else ""
        )

        form_fields[field_name] = form_field

    DynamicStageForm = type(
        "DynamicStageForm",
        (forms.Form,),
        form_fields
    )

    return (
        DynamicStageForm(data=data, files=files)
        if data is not None
        else DynamicStageForm(initial=initial)
    )
# =========================================================================================================== 
# ========================================== Stages =========================================================
# ===========================================================================================================



from .models import FormBatch, BatchPart

# @login_required
# def start_form(request, form_id):

#     form = get_object_or_404(Form, id=form_id)
#     first_stage = form.stages.order_by("order").first()

#     #  Safety check
#     if not first_stage:
#         messages.error(request, "No stages found for this form.")
#         return redirect("kanban_board", form_id=form.id)

#     if request.method == "POST":

#         batch_id = request.POST.get("batch_id")
#         company_name = request.POST.get("company_name")
#         part_id = request.POST.get("part_id")

#         if not batch_id or not company_name or not part_id:
#             messages.error(
#                 request,
#                 "Batch ID, Company Name and Part ID are required."
#             )
#             return redirect("kanban_board", form_id=form.id)

#         # -------------------------------------------------
#         # Create or get batch (WITH current_stage FIX)
#         # -------------------------------------------------
#         batch, created = FormBatch.objects.get_or_create(
#             form=form,
#             batch_id=batch_id,
#             defaults={
#                 "flow_id": uuid.uuid4(),
#                 "created_by": request.user,
#                 "company_name": company_name,
#                 "current_stage": first_stage,   #  REQUIRED FIX
#                 "is_active": True,
#             }
#         )

#         # -------------------------------------------------
#         # If batch already exists, ensure it has stage set
#         # (Extra safety for old records)
#         # -------------------------------------------------
#         if not batch.current_stage:
#             batch.current_stage = first_stage
#             batch.save(update_fields=["current_stage"])

#         # -------------------------------------------------
#         # Prevent duplicate Part ID inside same batch
#         # -------------------------------------------------
#         if batch.parts.filter(part_id=part_id).exists():
#             messages.error(request, "Part ID already exists in this Batch.")
#             return redirect("kanban_board", form_id=form.id)

#         # -------------------------------------------------
#         # Create new part
#         # -------------------------------------------------
#         part = BatchPart.objects.create(
#             batch=batch,
#             part_id=part_id,
#             current_stage=first_stage,
#             is_active=True
#         )

#         # -------------------------------------------------
#         # Redirect using part_id
#         # -------------------------------------------------
#         return redirect(
#             "fill_stage",
#             form_id=form.id,
#             stage_id=first_stage.id,
#             part_id=part.id
#         )

#     return redirect("kanban_board", form_id=form.id)


@login_required
def start_form(request, form_id):

    form = get_object_or_404(Form, id=form_id)
    first_stage = form.stages.order_by("order").first()

    if not first_stage:
        messages.error(request, "No stages found for this form.")
        return redirect("kanban_board", form_id=form.id)

    if request.method == "POST":

        batch_id = (request.POST.get("batch_id") or "").strip()
        company_name = (request.POST.get("company_name") or "").strip()
        part_id = (request.POST.get("part_id") or "").strip()
        material_batch_id = request.POST.get("material_batch")

        material_batch = None

        if material_batch_id:
            material_batch = MaterialBatch.objects.filter(id=material_batch_id).first()

        # -------------------------------------------------
        # REQUIRED FIELD VALIDATION
        # -------------------------------------------------

        if not batch_id:
            messages.error(request, "Batch ID is required.")
            return redirect("kanban_board", form_id=form.id)

        if not company_name:
            messages.error(request, "Company Name is required.")
            return redirect("kanban_board", form_id=form.id)

        # -------------------------------------------------
        # PART ID RULE
        # -------------------------------------------------

        if form.require_part_id:
            if not part_id:
                messages.error(request, "Part ID is required for this form.")
                return redirect("kanban_board", form_id=form.id)
        else:
            if not part_id:
                part_id = "-"

        try:
            with transaction.atomic():

                # -------------------------------------------------
                # STEP 1: CREATE OR GET BATCH
                # -------------------------------------------------

                batch, created = FormBatch.objects.get_or_create(
                    form=form,
                    batch_id=batch_id,
                    company_name=company_name,
                    defaults={
                        "flow_id": uuid.uuid4(),
                        "created_by": request.user,
                        "origin_user": request.user,   # traceability
                        "company_name": company_name,
                        "current_stage": first_stage,
                        "material_batch": material_batch,
                        "is_active": True
                    }
                )

                # -------------------------------------------------
                # UPDATE EXISTING BATCH
                # -------------------------------------------------

                if not created:

                    if not batch.is_active:
                        batch.is_active = True
                        batch.current_stage = first_stage

                    batch.company_name = company_name

                    if material_batch:
                        batch.material_batch = material_batch

                    batch.save(
                        update_fields=[
                            "is_active",
                            "current_stage",
                            "company_name",
                            "material_batch"
                        ]
                    )

                # -------------------------------------------------
                # STEP 2: GLOBAL PART ID CHECK
                # -------------------------------------------------

                existing_part = BatchPart.objects.filter(
                    part_id=part_id
                ).select_related("batch").first()

                if existing_part:

                    messages.error(
                        request,
                            f'⚠ Part ID "{part_id}" already exists in company '
                            f'"{existing_part.batch.company_name}" (Batch: {existing_part.batch.batch_id}).'
                        )

                    return redirect("kanban_board", form_id=form.id)

                # -------------------------------------------------
                # STEP 3: CREATE NEW PART
                # -------------------------------------------------

                part = BatchPart.objects.create(
                    batch=batch,
                    part_id=part_id,
                    current_stage=first_stage,
                    created_by=request.user,
                    moved_by=request.user,
                    status="ACTIVE",
                    is_active=True
                )
                AuditLog.objects.create(
                    user=request.user,
                    role=request.user.role,
                    module="Batch Creation",
                    action="CREATE",
                    model_name="BatchPart",
                    object_id=str(part.id),
                    object_repr=f"{batch.batch_id} - {part.part_id}",
                    new_data={
                        "folder": form.folder.name if form.folder else "-",
                        "form": form.name,
                        "stage": first_stage.name
                    },
                    ip_address=request.META.get("REMOTE_ADDR")
                )
                # -------------------------------------------------
                # REDIRECT TO FIRST STAGE FORM
                # -------------------------------------------------

                return redirect(
                    "fill_stage",
                    form_id=form.id,
                    stage_id=first_stage.id,
                    part_id=part.id
                )

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    return redirect("kanban_board", form_id=form.id)

# =============================================================================================================
# ======================================== Kanban Board =======================================================
# =============================================================================================================


@login_required
def kanban_board(request, form_id):

    form_obj = get_object_or_404(Form, id=form_id)

    # =====================================================
    # BASE QUERY (Admin vs User)
    # =====================================================

    if request.user.is_superuser or request.user.role == "admin":

        base_query = BatchPart.objects.filter(
            is_active=True
        )

        history_query = StageHistory.objects.filter(
            batch__form=form_obj
        )

    else:

        base_query = BatchPart.objects.filter(
            is_active=True,
            created_by=request.user
        )

        history_query = StageHistory.objects.filter(
            batch__form=form_obj,
            batch__parts__created_by=request.user
        ).distinct()

    # =====================================================
    # ACTIVE PARTS (NORMAL WORKFLOW)
    # =====================================================

    active_parts = base_query.filter(
        current_stage__form=form_obj,
        status="ACTIVE"
    ).select_related(
        "current_stage",
        "batch",
        "batch__material_batch",
        "created_by",
        "moved_by",
    ).order_by("-created_at")

    # =====================================================
    # NCR PARTS
    # =====================================================

    ncr_parts = base_query.filter(
        current_stage__form=form_obj,
        status="NCR"
    ).select_related(
        "current_stage",
        "batch",
        "batch__material_batch",
        "created_by",
        "moved_by",
    ).prefetch_related(
        "ncr_records"
    ).order_by("-created_at")

    # =====================================================
    # STAGE LIST
    # =====================================================

    stages = form_obj.stages.order_by("order")

    last_stage = stages.last()

    # Mark completion state for UI
    for part in active_parts:
        part.is_completed = (
            last_stage and part.current_stage_id == last_stage.id
        )

    # =====================================================
    # MATERIAL BATCH LIST (for modal dropdown)
    # =====================================================

    material_batches = MaterialBatch.objects.select_related(
        "created_by"
    ).order_by("-received_date")

    # =====================================================
    # AVAILABLE FORMS (FOR CONTINUE WORKFLOW)
    # =====================================================

    available_forms = Form.objects.exclude(id=form_obj.id)

    # =====================================================
    # STAGE HISTORY
    # =====================================================

    history = history_query.select_related(
        "from_stage",
        "to_stage",
        "moved_by",
        "batch"
    ).order_by("-moved_at")[:30]

    # =====================================================
    # OPTIONAL: STAGE COUNTERS (for dashboard)
    # =====================================================

    stage_counts = {}

    for stage in stages:
        stage_counts[stage.id] = active_parts.filter(
            current_stage=stage
        ).count()

    # =====================================================
    # RETURN CONTEXT
    # =====================================================

    return render(
        request,
        "qms_app/form/kanban_board.html",
        {
            "form_obj": form_obj,
            "stages": stages,
            "parts": active_parts,
            "ncr_parts": ncr_parts,
            "history": history,
            "available_forms": available_forms,
            "material_batches": material_batches,
            "stage_counts": stage_counts,
        }
    )

@login_required
def redirect_to_first_stage(request, form_id, batch_id):

    form_obj = get_object_or_404(Form, id=form_id)
    existing_batch = FormBatch.objects.filter(batch_id=batch_id).first()
    company_name = existing_batch.company_name if existing_batch else "UNKNOWN"

    #  Check if batch already exists for this form
    batch = FormBatch.objects.filter(
        form=form_obj,
        batch_id=batch_id
    ).first()

    if not batch:
        #  Create new batch for this form
        first_stage = form_obj.stages.order_by("order").first()

        if not first_stage:
            messages.error(request, "No stages found in this form.")
            return redirect("kanban_board", form_id=form_id)

        batch = FormBatch.objects.create(
            form=form_obj,
            batch_id=batch_id,
            company_name=company_name,
            current_stage=first_stage,
            created_by=request.user
        )

    #  Redirect directly to first stage
    return redirect(
        "fill_stage",
        form_id=form_id,
        stage_id=batch.current_stage.id,
        batch_id=batch.batch_id
    )



@login_required
def continue_to_next_form(request, form_id, part_id):

    form_obj = get_object_or_404(Form, id=form_id)

    previous_part = BatchPart.objects.select_related("batch").filter(id=part_id).first()

    if not previous_part:
        messages.error(request, "Previous part not found.")
        return redirect("form_list")

    previous_batch = previous_part.batch

    # FIRST STAGE OF NEXT FORM
    first_stage = form_obj.stages.order_by("order").first()

    if not first_stage:
        messages.error(request, "No stages found in this form.")
        return redirect("kanban_board", form_id=form_id)

    with transaction.atomic():

        # CREATE OR GET BATCH FOR NEXT FORM
        new_batch, created = FormBatch.objects.get_or_create(
            form=form_obj,
            batch_id=previous_batch.batch_id,
            company_name=previous_batch.company_name,
            defaults={
                "flow_id": previous_batch.flow_id,
                "current_stage": first_stage,
                "created_by": previous_batch.created_by,
                "company_name": previous_batch.company_name,
                "material_batch": previous_batch.material_batch,
                "is_active": True
            }
        )

        # ACTIVATE BATCH IF NEEDED
        if not new_batch.is_active:
            new_batch.is_active = True
            new_batch.current_stage = first_stage
            new_batch.save(update_fields=["is_active", "current_stage"])

        old_stage = previous_part.current_stage

        # MOVE PART TO NEW FORM
        previous_part.batch = new_batch
        previous_part.current_stage = first_stage
        previous_part.moved_by = request.user
        previous_part.is_active = True

        previous_part.save(update_fields=[
            "batch",
            "current_stage",
            "moved_by",
            "is_active"
        ])

        # SAVE STAGE HISTORY
        StageHistory.objects.create(
            batch=new_batch,
            part=previous_part,
            from_stage=old_stage,
            to_stage=first_stage,
            moved_by=request.user
        )

        # AUDIT LOG
        AuditLog.objects.create(
            user=request.user,
            role=request.user.role,
            module="Workflow",
            action="FORM_MOVED",
            model_name="BatchPart",
            object_id=str(previous_part.id),
            object_repr=f"{new_batch.batch_id} - {previous_part.part_id}",
            old_data={
                "folder": previous_batch.form.folder.name if previous_batch.form.folder else "-",
                "form": previous_batch.form.name,
                "stage": old_stage.name if old_stage else None
            },
            new_data={
                "folder": form_obj.folder.name if form_obj.folder else "-",
                "form": form_obj.name,
                "stage": first_stage.name
            },
            ip_address=request.META.get("REMOTE_ADDR")
        )

        # DEACTIVATE OLD BATCH IF EMPTY
        if not previous_batch.parts.exclude(id=previous_part.id).filter(is_active=True).exists():
            previous_batch.is_active = False
            previous_batch.save(update_fields=["is_active"])

    return redirect("kanban_board", form_id=form_obj.id)



# from django.views.decorators.http import require_POST
# from django.urls import reverse
# from .models import BatchPart, Stage, StageHistory

# @require_POST
# @login_required
# def move_kanban_card(request):

#     data = json.loads(request.body)

#     part_id = data.get("part_id")   #  changed
#     target_stage_id = data.get("stage_id")

#     if not part_id or not target_stage_id:
#         return JsonResponse(
#             {"error": "Invalid data"},
#             status=400
#         )

#     #  Get Part (NOT batch)
#     part = get_object_or_404(BatchPart, id=part_id)
#     target_stage = get_object_or_404(Stage, id=target_stage_id)

#     current_stage = part.current_stage

#     #  RULE 1: Only NEXT stage allowed
#     if target_stage.order != current_stage.order + 1:
#         return JsonResponse(
#             {"error": "You can only move to the next stage"},
#             status=403
#         )

#     #  RULE 2: Current stage MUST be completed
#     is_completed = FormSubmission.objects.filter(
#         stage=current_stage,
#         part=part   #  changed from submission_batch
#     ).exists()

#     if not is_completed:
#         return JsonResponse(
#             {"error": "Complete this stage before moving forward"},
#             status=403
#         )

#     #  SAVE STAGE HISTORY (still linked to batch for grouping)
#     StageHistory.objects.create(
#         batch=part.batch,
#         from_stage=current_stage,
#         to_stage=target_stage,
#         moved_by=request.user
#     )

#     #  MOVE PART
#     part.current_stage = target_stage
#     part.save(update_fields=["current_stage"])

#     #  AUDIT LOG
#     AuditLog.objects.create(
#         user=request.user,
#         role=request.user.role,
#         module="Form Workflow",
#         action="STAGE_MOVED",
#         model_name="BatchPart",
#         object_repr=f"{part.batch.batch_id} - {part.part_id}",
#         old_data={
#             "stage": current_stage.name,
#             "form":  current_stage.form.name
#         },
#         new_data={
#             "stage": target_stage.name,
#             "part_id": part.part_id,
#             "form": target_stage.form.name,
#             "batch_id": part.batch.batch_id
#         },
#         ip_address=request.META.get("REMOTE_ADDR")
#     )

#     return JsonResponse({
#         "success": True,
#         "redirect_url": reverse(
#             "fill_stage",
#             args=[
#                 part.batch.form.id,
#                 target_stage.id,
#                 part.id    #  redirect using part_id
#             ]
#         )
#     })

@login_required
def move_kanban_card(request):

    data = json.loads(request.body)

    part_id = data.get("part_id")
    target_stage_id = data.get("stage_id")

    if not part_id or not target_stage_id:
        return JsonResponse(
            {"error": "Invalid data"},
            status=400
        )

    # Get Part
    part = get_object_or_404(BatchPart, id=part_id)
    target_stage = get_object_or_404(Stage, id=target_stage_id)

    current_stage = part.current_stage

    # RULE 1: Only NEXT stage allowed
    if target_stage.order != current_stage.order + 1:
        return JsonResponse(
            {"error": "You can only move to the next stage"},
            status=403
        )

    # RULE 2: Current stage MUST be completed
    is_completed = FormSubmission.objects.filter(
        stage=current_stage,
        part=part
    ).exists()

    if not is_completed:
        return JsonResponse(
            {"error": "Complete this stage before moving forward"},
            status=403
        )

    # SAVE STAGE HISTORY
    StageHistory.objects.create(
        batch=part.batch,
        part=part,
        from_stage=current_stage,
        to_stage=target_stage,
        moved_by=request.user
    )

    # MOVE PART
    part.current_stage = target_stage
    part.save(update_fields=["current_stage"])

    # AUDIT LOG
    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="Form Workflow",
        action="STAGE_MOVED",
        model_name="BatchPart",
        object_id=str(part.id),
        object_repr=f"{part.batch.batch_id} - {part.part_id}",
        old_data={
            "folder": current_stage.form.folder.name if current_stage.form.folder else "-",
            "stage": current_stage.name,
            "form": current_stage.form.name,
        },
        new_data={
            "folder": target_stage.form.folder.name if target_stage.form.folder else "-",
            "stage": target_stage.name,
            "form": target_stage.form.name,
        },
        ip_address=request.META.get("REMOTE_ADDR")
    )

    return JsonResponse({
        "success": True,
        "redirect_url": reverse(
            "fill_stage",
            args=[
                part.batch.form.id,
                target_stage.id,
                part.id
            ]
        )
    })

from django.db import transaction
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.urls import reverse
import json


@login_required
def delete_form_batch(request):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    if request.user.role != "admin" and not request.user.is_superuser:
        return JsonResponse({"error": "Only admin can delete parts."}, status=403)

    # -----------------------------------------------------
    # Support BOTH JSON (Kanban) and FORM POST
    # -----------------------------------------------------

    if request.content_type == "application/json":
        try:
            data = json.loads(request.body)
            part_id = data.get("part_id")
            action = data.get("action")
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
    else:
        part_id = request.POST.get("part_id")
        action = request.POST.get("action", "delete")

    if not part_id:
        return JsonResponse({"error": "Part ID required"}, status=400)

    try:
        with transaction.atomic():

            part = BatchPart.objects.select_related(
                "batch",
                "batch__form",
                "current_stage"
            ).get(id=part_id)

            batch = part.batch
            flow_id = batch.flow_id
            part_code = part.part_id
            signature_key = f"{flow_id}_{part_code}"

            # =====================================================
            # MOVE BACK TO PREVIOUS FORM
            # =====================================================
            if action == "previous":

                previous_history = (
                    StageHistory.objects
                    .filter(part=part)
                    .select_related("from_stage", "batch")
                    .order_by("-moved_at")
                    .first()
                )

                if not previous_history:
                    return JsonResponse({
                        "success": False,
                        "error": "No previous form found."
                    })

                previous_stage = previous_history.from_stage
                previous_form = previous_stage.form

                # ---------------- Audit Log ----------------
                AuditLog.objects.create(
                    user=request.user,
                    role=request.user.role,
                    module="Form Workflow",
                    action="MOVE_BACK",
                    model_name="BatchPart",
                    object_id=str(part.id),
                    object_repr=f"{batch.batch_id} - {part_code}",
                    old_data={
                        "form": batch.form.name,
                        "stage": part.current_stage.name if part.current_stage else "-"
                    },
                    new_data={
                        "form": previous_form.name,
                        "stage": previous_stage.name
                    },
                    ip_address=request.META.get("REMOTE_ADDR")
                )

                # Move card back
                part.current_stage = previous_stage
                part.batch.form = previous_form
                part.save(update_fields=["current_stage"])

                # Remove current stage submission
                FormSubmission.objects.filter(
                    part=part,
                    stage=batch.current_stage
                ).delete()

                # Remove latest stage history entry
                previous_history.delete()

                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("kanban_board", args=[previous_form.id])
                })

            # =====================================================
            # COMPLETE HARD DELETE
            # =====================================================
            elif action == "delete":

                related_batches = FormBatch.objects.filter(flow_id=flow_id)

                related_parts = BatchPart.objects.filter(
                    batch__in=related_batches,
                    part_id=part_code
                )

                # ---------------- Audit BEFORE delete ----------------
                AuditLog.objects.create(
                    user=request.user,
                    role=request.user.role,
                    module="Form Workflow",
                    action="DELETE",
                    model_name="BatchPart",
                    object_id=str(part.id),
                    object_repr=f"{batch.batch_id} - {part_code}",
                    old_data={
                        "form": batch.form.name,
                        "stage": part.current_stage.name if part.current_stage else "-"
                    },
                    ip_address=request.META.get("REMOTE_ADDR")
                )

                # Delete submissions
                FormSubmission.objects.filter(
                    part__in=related_parts
                ).delete()

                # Delete signatures
                SignatureVerification.objects.filter(
                    batch_id__startswith=str(flow_id)
                ).delete()

                # Delete stage history
                StageHistory.objects.filter(
                    batch__in=related_batches
                ).delete()

                # Delete parts
                related_parts.delete()

                # Delete empty batches
                for b in related_batches:
                    if not b.parts.exists():
                        b.delete()

                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("form_list")
                })

            else:
                return JsonResponse({
                    "success": False,
                    "error": "Invalid action"
                })

    except BatchPart.DoesNotExist:
        return JsonResponse({"error": "Part not found"}, status=404)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



# ===================================================================================================================
# =========================================== fill Stage ============================================================
# ===================================================================================================================


# @login_required
# def fill_stage(request, form_id, stage_id, part_id=None):

#     form_obj = get_object_or_404(Form, id=form_id)
#     stage = get_object_or_404(Stage, id=stage_id)

#     prev_stage = form_obj.stages.filter(order__lt=stage.order).last()
#     next_stage = form_obj.stages.filter(order__gt=stage.order).first()

#     # ------------------------------------------------
#     # Get current PART object
#     # ------------------------------------------------
#     part_obj = None
#     if part_id:
#         part_obj = get_object_or_404(
#             BatchPart,
#             id=part_id,
#             batch__form=form_obj
#         )

#     # ------------------------------------------------
#     # Existing submission for this stage
#     # ------------------------------------------------
#     existing_submission = None
#     if part_obj:
#         existing_submission = FormSubmission.objects.filter(
#             form=form_obj,
#             stage=stage,
#             part=part_obj
#         ).first()

#     # ------------------------------------------------
#     # Previous stage completion check
#     # ------------------------------------------------
#     prev_submission = None
#     if prev_stage and part_obj:
#         prev_submission = FormSubmission.objects.filter(
#             form=form_obj,
#             stage=prev_stage,
#             part=part_obj
#         ).first()

#     role_allowed = stage.can_be_filled_by(request.user)
#     prev_completed = True if not prev_stage else bool(prev_submission)
#     can_submit = role_allowed and prev_completed

#     if request.method == "POST" and not can_submit:
#         return HttpResponseForbidden("Submission not allowed.")

#     # ==========================================================
#     #  PREFILL LOGIC (Same Form + Cross Form)
#     # ==========================================================
#     prefill_initial = {}

#     if part_obj:

#         label_type_map = {}

#         # ------------------------------------------------------
#         #  SAME FORM - Previous Stage Submissions
#         # ------------------------------------------------------
#         same_form_previous_submissions = FormSubmission.objects.filter(
#             form=form_obj,
#             part=part_obj,
#             stage__order__lt=stage.order
#         )

#         # ------------------------------------------------------
#         #  OTHER FORMS - Same Flow
#         # ------------------------------------------------------
#         related_batches = FormBatch.objects.filter(
#             flow_id=part_obj.batch.flow_id
#         ).exclude(form=form_obj)

#         related_parts = BatchPart.objects.filter(
#             batch__in=related_batches,
#             part_id=part_obj.part_id
#         )

#         cross_form_submissions = FormSubmission.objects.filter(
#             part__in=related_parts
#         )

#         # ------------------------------------------------------
#         #  Combine both submission sources
#         # ------------------------------------------------------
#         all_previous_submissions = list(same_form_previous_submissions) + list(cross_form_submissions)

#         for submission in all_previous_submissions:
#             for key, value in submission.data.items():

#                 if not key.startswith("field_"):
#                     continue

#                 field_id = key.replace("field_", "")
#                 field_obj = FormField.objects.filter(id=field_id).first()
#                 if not field_obj:
#                     continue

#                 label = (field_obj.label or "").strip().lower()
#                 field_type = field_obj.field_type

#                 if (label, field_type) not in label_type_map and value:
#                     label_type_map[(label, field_type)] = value

#         # ------------------------------------------------------
#         # Apply matching label + type
#         # ------------------------------------------------------
#         for field in stage.fields.all():
#             key = f"field_{field.id}"
#             label = (field.label or "").strip().lower()
#             field_type = field.field_type

#             if (label, field_type) in label_type_map:
#                 prefill_initial[key] = label_type_map[(label, field_type)]

#         # ------------------------------------------------------
#         # Auto Prefill Part ID
#         # ------------------------------------------------------
#         for field in stage.fields.all():
#             if (field.label or "").strip().lower() == "part id":
#                 prefill_initial[f"field_{field.id}"] = part_obj.part_id

#     # -------------------------
#     # Build Form (Keep Your Merge Logic)
#     # -------------------------
#     if request.method == "POST":
#         form = generate_form(stage, data=request.POST, files=request.FILES)
#     else:
#         if existing_submission:
#             combined_initial = existing_submission.data.copy()
#             combined_initial.update(prefill_initial)
#             form = generate_form(stage, initial=combined_initial)
#         else:
#             form = generate_form(stage, initial=prefill_initial)

#     # --------------------------------------------------------
#     # Lock Part ID Field
#     # --------------------------------------------------------
#     if part_obj:
#         for name, field in form.fields.items():
#             if (field.label or "").strip().lower() == "part id":
#                 field.initial = part_obj.part_id
#                 field.disabled = True

#     if not can_submit:
#         for f in form.fields.values():
#             f.disabled = True

#     # -------------------------
#     # SAVE (UNCHANGED)
#     # -------------------------
#     if request.method == "POST" and form.is_valid():

#         submission_data = {}

#         for name, value in form.cleaned_data.items():
#             field = form.fields.get(name)

#             if field.widget.attrs.get("is_table"):
#                 raw_json = request.POST.get(name)
#                 try:
#                     parsed = json.loads(raw_json) if raw_json else []

#                     #  Validate digital signature cells safely
#                     for row in parsed:
#                         for cell in row.get("cells", []):

#                             if cell.get("type") == "signature":

#                                 value = cell.get("value")

#                                 # Ensure value is dictionary
#                                 if not isinstance(value, dict):
#                                     cell["value"] = {}

#                                 else:
#                                 # Ensure required keys exist
#                                     cell["value"] = {
#                                         "signed_by": value.get("signed_by"),
#                                         "signed_at": value.get("signed_at"),
#                                         "reference_code": value.get("reference_code")
#                                     }

#                         submission_data[name] = parsed

#                 except json.JSONDecodeError:
#                     submission_data[name] = []

#             elif hasattr(value, "name"):
#                 file_path = default_storage.save(
#                     f"stage_uploads/{value.name}",
#                     ContentFile(value.read())
#                 )
#                 submission_data[name] = file_path

#             else:
#                 if isinstance(value, (date, datetime)):
#                     submission_data[name] = value.isoformat()
#                 else:
#                     submission_data[name] = value

#         if existing_submission:
#             existing_submission.data = submission_data
#             existing_submission.submitted_by = request.user
#             existing_submission.submitted_at = timezone.now()
#             existing_submission.save()
#         else:
#             FormSubmission.objects.create(
#                 form=form_obj,
#                 stage=stage,
#                 submitted_by=request.user,
#                 submitted_at=timezone.now(),
#                 data=submission_data,
#                 part=part_obj
#             )

#             readable_data = {}

#             for key, value in submission_data.items():
#                 if key.startswith("field_"):
#                     field_id = key.replace("field_", "")
#                     field_obj = FormField.objects.filter(id=field_id).first()

#                     if field_obj:
#                         readable_data[field_obj.label] = value
#                     else:
#                         readable_data[key] = value
#                 else:
#                     readable_data[key] = value

#             AuditLog.objects.create(
#                 user=request.user,
#                 role=request.user.role,
#                 module="FormSubmission",
#                 action="CREATE",
#                 model_name="FormSubmission",
#                 object_repr=f"{part_obj.batch.batch_id} - {part_obj.part_id}",
#                 new_data={
#                     "form": form_obj.name,
#                     "stage": stage.name,
#                     "data": readable_data
#                 },
#                 ip_address=request.META.get("REMOTE_ADDR")
#             )

#         messages.success(request, "Stage submitted successfully.")
#         return redirect("kanban_board", form_id=form_obj.id)

#     return render(
#         request,
#         "qms_app/form/fill_stage.html",
#         {
#             "form_obj": form_obj,
#             "stage": stage,
#             "form": form,
#             "part_id": part_id,
#             "existing_submission": existing_submission,
#             "prev_stage": prev_stage,
#             "next_stage": next_stage,
#             "can_submit": can_submit,
#         }
#     )

@login_required
def fill_stage(request, form_id, stage_id, part_id=None):

    form_obj = get_object_or_404(Form, id=form_id)
    stage = get_object_or_404(Stage, id=stage_id)

    prev_stage = form_obj.stages.filter(order__lt=stage.order).last()
    next_stage = form_obj.stages.filter(order__gt=stage.order).first()

    # ------------------------------------------------
    # Get current PART object
    # ------------------------------------------------
    part_obj = None
    if part_id:
        part_obj = get_object_or_404(
            BatchPart,
            id=part_id,
            batch__form=form_obj
        )

    if part_obj:
        request.session["current_signature_key"] = str(part_obj.id)

    # ------------------------------------------------
    # Existing submission for this stage
    # ------------------------------------------------
    existing_submission = None
    if part_obj:
        existing_submission = FormSubmission.objects.filter(
            form=form_obj,
            stage=stage,
            part=part_obj
        ).first()

    # ------------------------------------------------
    # Previous stage completion check
    # ------------------------------------------------
    prev_submission = None
    if prev_stage and part_obj:
        prev_submission = FormSubmission.objects.filter(
            form=form_obj,
            stage=prev_stage,
            part=part_obj
        ).first()

    role_allowed = stage.can_be_filled_by(request.user)
    prev_completed = True if not prev_stage else bool(prev_submission)
    can_submit = role_allowed and prev_completed

    if request.method == "POST" and not can_submit:
        return HttpResponseForbidden("Submission not allowed.")

    # ==========================================================
    # PREFILL LOGIC (UNCHANGED)
    # ==========================================================
    prefill_initial = {}

    if part_obj:

        label_type_map = {}

        same_form_previous_submissions = FormSubmission.objects.filter(
            form=form_obj,
            part=part_obj,
            stage__order__lt=stage.order
        )

        related_batches = FormBatch.objects.filter(
            flow_id=part_obj.batch.flow_id
        ).exclude(form=form_obj)

        related_parts = BatchPart.objects.filter(
            batch__in=related_batches,
            part_id=part_obj.part_id
        )

        cross_form_submissions = FormSubmission.objects.filter(
            part__in=related_parts
        )

        all_previous_submissions = list(same_form_previous_submissions) + list(cross_form_submissions)

        for submission in all_previous_submissions:
            for key, value in submission.data.items():

                if not key.startswith("field_"):
                    continue

                field_id = key.replace("field_", "")
                field_obj = FormField.objects.filter(id=field_id).first()
                if not field_obj:
                    continue

                label = (field_obj.label or "").strip().lower()
                field_type = field_obj.field_type

                if (label, field_type) not in label_type_map and value:
                    label_type_map[(label, field_type)] = value

        for field in stage.fields.all():
            key = f"field_{field.id}"
            label = (field.label or "").strip().lower()
            field_type = field.field_type

            if (label, field_type) in label_type_map:
                prefill_initial[key] = label_type_map[(label, field_type)]

        #  Always prefill Part ID from BatchPart
        for field in stage.fields.all():
            if (field.label or "").strip().lower() == "part id":
                prefill_initial[f"field_{field.id}"] = part_obj.part_id

    # -------------------------
    # Build Form
    # -------------------------
    if request.method == "POST":
        form = generate_form(stage, data=request.POST, files=request.FILES)
    else:
        if existing_submission:
            combined_initial = existing_submission.data.copy()
            combined_initial.update(prefill_initial)
            form = generate_form(stage, initial=combined_initial)
        else:
            form = generate_form(stage, initial=prefill_initial)
    # ==========================================================
# 🔥 LOAD TABLE DATA INTO WIDGET (CRITICAL FIX)
# ==========================================================

    if request.method != "POST" and existing_submission:
        for name, field in form.fields.items():

            if field.widget.attrs.get("is_table"):

                saved_rows = existing_submission.data.get(name, [])

                if isinstance(saved_rows, str):
                    try:
                        saved_rows = json.loads(saved_rows)
                    except:
                        saved_rows = []

                field.widget.attrs['rows'] = saved_rows
    # --------------------------------------------------------
    # Lock Part ID
    # --------------------------------------------------------
    if part_obj:
        for name, field in form.fields.items():
            if (field.label or "").strip().lower() == "part id":
                field.initial = part_obj.part_id
                if part_obj.part_id and part_obj.part_id != "-":
                    field.disabled = True

    if not can_submit:
        for f in form.fields.values():
            f.disabled = True

    # ==========================================================
    # SAVE (MERGE TABLE DATA + RESTORE PART ID SYNC)
    # ==========================================================
    if request.method == "POST" and form.is_valid():

        submission_data = {}

        for name, value in form.cleaned_data.items():
            field = form.fields.get(name)

            # ---------------- TABLE FIELD ----------------
            if field.widget.attrs.get("is_table"):

                raw_json = request.POST.get(name)

                try:
                    new_data = json.loads(raw_json) if raw_json else []

                    for row in new_data:
                        for cell in row.get("cells", []):

                            if cell.get("type") == "signature":

                                sig_value = cell.get("value", {})

                                if isinstance(sig_value, dict):

                                    signed_by = sig_value.get("signed_by")

                                    if signed_by and "@" in signed_by:

                                        user = User.objects.filter(email=signed_by).first()

                                        if user:
                                            sig_value["signed_by"] = user.get_display_name()

                      
                                    cell["value"] = sig_value

                    submission_data[name] = new_data

                except json.JSONDecodeError:
                    submission_data[name] = []

            # ---------------- FILE FIELD ----------------
            elif hasattr(value, "name"):
                file_path = default_storage.save(
                    f"stage_uploads/{value.name}",
                    ContentFile(value.read())
                )
                submission_data[name] = file_path

            # ---------------- NORMAL FIELD ----------------
            else:
                if isinstance(value, (date, datetime)):
                    submission_data[name] = value.isoformat()
                else:
                    submission_data[name] = value

        # ==================================================
        #  CRITICAL FIX — SYNC PART ID TO BatchPart
        # ==================================================
        if part_obj:
            for name, value in submission_data.items():
                if name.startswith("field_"):

                    field_id = int(name.replace("field_", ""))
                    field_obj = FormField.objects.filter(id=field_id).first()

                    if field_obj and (field_obj.label or "").strip().lower() == "part id":

                        new_part_id = (value or "").strip()

                        if new_part_id and new_part_id != "-":

                            if not BatchPart.objects.filter(
                                batch=part_obj.batch,
                                part_id=new_part_id
                            ).exclude(id=part_obj.id).exists():

                                part_obj.part_id = new_part_id
                                part_obj.save(update_fields=["part_id"])
        # -------------------------
        # UPDATE EXISTING SUBMISSION
        # -------------------------

        if existing_submission:

            old_stage_name = "-"

            if existing_submission.stage:

                old_stage_name = (
                    existing_submission.stage.name
                )

            # =====================================
            # GET OLD SAVED DATA
            # =====================================

            old_data = (
                existing_submission.data
                or {}
            )

            # =====================================
            # COPY OLD DATA
            # =====================================

            merged_data = old_data.copy()

            # =====================================
            # MERGE NEW DATA
            # =====================================

            for key, value in submission_data.items():

                # =================================
                # TABLE FIELD
                # =================================

                if isinstance(value, list):

                    old_table = old_data.get(
                        key,
                        []
                    )

                    merged_rows = []

                    # =============================
                    # OLD ROW MAP
                    # =============================

                    old_row_map = {

                        str(row.get("_row_id")): row

                        for row in old_table

                        if isinstance(row, dict)
                    }

                    current_row_ids = set()

                    # =============================
                    # LOOP NEW ROWS
                    # =============================

                    for new_row in value:

                        if not isinstance(new_row, dict):

                            merged_rows.append(new_row)

                            continue

                        row_id = str(
                            new_row.get("_row_id")
                        )

                        current_row_ids.add(row_id)

                        old_row = old_row_map.get(
                            row_id,
                            {}
                        )

                        merged_row = old_row.copy()

                        # =========================
                        # MERGE CELLS
                        # =========================

                        for col, cell_value in new_row.items():

                            # preserve row id
                            if col == "_row_id":

                                merged_row[col] = cell_value

                                continue

                            # preserve old value
                            if cell_value in [
                                None,
                                ""
                            ]:

                                continue

                            # update new value
                            merged_row[col] = cell_value

                        merged_rows.append(
                            merged_row
                        )

                    # =============================
                    # REMOVE DELETED ROWS
                    # =============================

                    final_rows = []

                    for row in merged_rows:

                        row_id = str(
                            row.get("_row_id")
                        )

                        if row_id in current_row_ids:

                            final_rows.append(row)

                    merged_data[key] = final_rows

                # =================================
                # FILE FIELD
                # =================================

                elif isinstance(value, str) and (

                    value.startswith(
                        "stage_uploads/"
                    )

                ):

                    merged_data[key] = value

                # =================================
                # NORMAL FIELD
                # =================================

                else:

                    if value not in [
                        None,
                        "",
                        []
                    ]:

                        merged_data[key] = value
            # =====================================
            # SAVE FINAL MERGED DATA
            # =====================================

            existing_submission.data = merged_data

            existing_submission.submitted_by = (
                request.user
            )

            existing_submission.submitted_at = (
                timezone.now()
            )

            existing_submission.save()

            # =====================================
            # AUDIT LOG
            # =====================================

            AuditLog.objects.create(

                user=request.user,

                role=request.user.role,

                module="Form Submission",

                action="STAGE_SUBMITTED",

                model_name="FormSubmission",

                object_id=str(part_obj.id),

                object_repr=f"{part_obj.batch.batch_id} - {part_obj.part_id}",

                old_data={

                    "form": form_obj.name,

                    "stage": old_stage_name
                },

                new_data={

                    "folder": (
                        form_obj.folder.name
                        if form_obj.folder
                        else "-"
                    ),

                    "form": form_obj.name,

                    "stage": stage.name
                },

                ip_address=request.META.get(
                    "REMOTE_ADDR"
                )
            )

            messages.success(
                request,
                "Stage updated successfully."
            )

            return redirect(
                "kanban_board",
                form_id=form_obj.id
            )
            
        else:
            FormSubmission.objects.create(
                form=form_obj,
                stage=stage,
                submitted_by=request.user,
                submitted_at=timezone.now(),
                data=submission_data,
                part=part_obj
            )

            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="Form Submission",
                action="STAGE_SUBMITTED",
                model_name="FormSubmission",
                object_id=str(part_obj.id),
                object_repr=f"{part_obj.batch.batch_id} - {part_obj.part_id}",
                new_data={
                    "folder": form_obj.folder.name if form_obj.folder else "-",
                    "form": form_obj.name,
                    "stage": stage.name
                },
                ip_address=request.META.get("REMOTE_ADDR")
            )

            messages.success(request, "Stage submitted successfully.")
            return redirect("kanban_board", form_id=form_obj.id)

    return render(
        request,
        "qms_app/form/fill_stage.html",
        {
            "form_obj": form_obj,
            "stage": stage,
            "form": form,
            "part_id": part_id,
            "existing_submission": existing_submission,
            "prev_stage": prev_stage,
            "next_stage": next_stage,
            "can_submit": can_submit,
        }
    )


@login_required
def workflow_list(request):
    """
    Shows all workflow batches.
    Every user can see all batches.
    """
    batches = (
        FormSubmission.objects
        .values("form_id", "submission_batch", "form__name")
        .distinct()
        .order_by("-submission_batch")
    )

    return render(
        request,
        "qms_app/workflow/workflow_list.html",
        {"batches": batches},
    )


@login_required
def view_form_batch(request, form_id, batch_id):
    form_obj = get_object_or_404(Form, id=form_id)

    submissions = (
        FormSubmission.objects
        .filter(form=form_obj, submission_batch=batch_id)
        .select_related("stage", "submitted_by")
        .order_by("stage__order")
    )

    if not submissions.exists():
        raise Http404("Invalid or empty batch")

    submission_map = {s.stage_id: s for s in submissions}

    stages_info = []
    previous_completed = True

    for stage in form_obj.stages.all().order_by("order"):
        submission = submission_map.get(stage.id)
        is_completed = submission is not None

        can_fill = (
            stage.can_be_filled_by(request.user)
            and previous_completed
            and not is_completed
        )

        #  NEW FLAG
        can_edit_completed = (
            is_completed and stage.can_be_filled_by(request.user)
        )

        stages_info.append({
            "stage": stage,
            "is_completed": is_completed,
            "can_fill": can_fill,
            "is_current": can_fill,

            "submitted_by": submission.submitted_by if submission else None,
            "submitted_at": submission.submitted_at if submission else None,

            "allowed_roles": stage.allowed_roles,
            "can_edit_completed": can_edit_completed,  #  IMPORTANT
        })

        if not is_completed:
            previous_completed = False

    return render(
        request,
        "qms_app/form/view_batch.html",
        {
            "form_obj": form_obj,
            "batch_id": batch_id,
            "stages_info": stages_info,
        }
    )

@login_required
def view_stage_data(request, stage_id, batch_id):
    stage = get_object_or_404(Stage, id=stage_id)

    submission = get_object_or_404(
        FormSubmission,
        stage=stage,
        submission_batch=batch_id
    )

    #  Build label → value pairs
    display_data = []

    for key, value in submission.data.items():
        label = key

        if key.startswith("field_"):
            try:
                field_id = int(key.replace("field_", ""))
                field_obj = FormField.objects.get(id=field_id)
                label = field_obj.label
            except FormField.DoesNotExist:
                label = key

        # Handle multi-select nicely
        if isinstance(value, list):
            value = ", ".join(value)

        display_data.append({
            "label": label,
            "value": value,
        })

    return render(
        request,
        "qms_app/form/view_stage_data.html",
        {
            "stage": stage,
            "submission": submission,
            "batch_id": batch_id,
            "display_data": display_data,
        },
    )






# =======================================================================================================================
# =============================================== Form List =============================================================
# =======================================================================================================================

from django.db.models import Max
@require_page_permission("can_submitted")
@login_required
def submitted_forms_list(request):

    # --------------------------------------------------
    # Get unique workflow IDs from FormBatch
    # --------------------------------------------------
    flows = (
        FormBatch.objects
        .values_list("flow_id", flat=True)
        .distinct()
    )

    form_data = []

    for flow_id in flows:

        # Get batches under this workflow
        flow_batches = FormBatch.objects.filter(flow_id=flow_id)

        if not flow_batches.exists():
            continue

        # Get parts under those batches
        parts = BatchPart.objects.filter(batch__in=flow_batches)

        if not parts.exists():
            continue

        # Check if submissions exist for those parts
        has_submission = FormSubmission.objects.filter(
            part__in=parts
        ).exists()

        if not has_submission:
            continue

        first_batch = flow_batches.first()

        form_data.append({
            "flow_id": flow_id,
            "company_name": first_batch.company_name,
            "batch_id": first_batch.batch_id,  # keep original batch id
            "created_by": first_batch.created_by.email,
            "created_at": first_batch.created_at,
            "forms_count": flow_batches.count(),
            "parts_count": parts.count(),  #  NEW (optional info)
        })

    # Optional: sort newest first
    form_data = sorted(form_data, key=lambda x: x["created_at"], reverse=True)

    return render(
        request,
        "qms_app/form/submitted_forms_list.html",
        {"form_data": form_data}
    )





@login_required
def verify_signature(request):

    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    password = data.get("password", "").strip()
    slot = data.get("slot")

    if slot not in ["prepared", "approved"]:
        return JsonResponse({"ok": False, "error": "Invalid slot"}, status=400)

    if not request.user.check_password(password):
        return JsonResponse({"ok": False, "error": "Invalid password"}, status=401)

    user = request.user

    signature_key = request.session.get("current_signature_key")

    if not signature_key:
        return JsonResponse({
            "ok": False,
            "error": "Signature key missing. Reload page."
        }, status=400)

    # ---------------------------------------------------------
    # Generate display name
    # ---------------------------------------------------------
    display_name = f"{user.first_name}{user.last_name}".strip()
    if not display_name:
        display_name = user.email.split("@")[0]
    # ---------------------------------------------------------
    # Generate unique reference code
    # ---------------------------------------------------------
    while True:
        ref_code = ''.join(
            random.choices(string.ascii_uppercase + string.digits, k=8)
        )
        if not SignatureVerification.objects.filter(reference_code=ref_code).exists():
            break

    # ---------------------------------------------------------
    # Create NEW signature every time
    # ---------------------------------------------------------
    sig = SignatureVerification.objects.create(
        user=user,
        display_name=display_name,
        reference_code=ref_code,
        ip_address=request.META.get("REMOTE_ADDR"),
        batch_id=str(signature_key),
        slot=slot,
    )

    return JsonResponse({
        "ok": True,
        "user": display_name,
        "time": timezone.localtime(sig.verified_at).strftime("%d %b %Y %H:%M"),
        "code": ref_code
    })
# @login_required
# @csrf_exempt
# def verify_signature(request):

#     if request.method != "POST":
#         return JsonResponse({"ok": False}, status=405)

#     try:
#         data = json.loads(request.body)
#     except Exception:
#         return JsonResponse({"ok": False}, status=400)

#     password = data.get("password", "").strip()
#     slot = data.get("slot")

#     if slot not in ["prepared", "approved"]:
#         return JsonResponse({"ok": False})

#     #  Authenticate user
#     user = authenticate(request, email=request.user.email, password=password)
#     if not user:
#         return JsonResponse({"ok": False})

#     #  IMPORTANT FIX:
#     # Use UNIQUE signature key per FLOW + PART
#     signature_key = request.session.get("current_signature_key")

#     if not signature_key:
#         return JsonResponse({
#             "ok": False,
#             "error": "Signature key missing"
#         }, status=400)

#     #  Check if already signed (per PART)
#     existing = SignatureVerification.objects.filter(
#         batch_id=signature_key,
#         slot=slot
#     ).first()

#     if existing:
#         return JsonResponse({
#             "ok": True,
#             "name": existing.display_name,
#             "time": timezone.localtime(existing.verified_at).strftime("%d %b %Y %H:%M"),
#             "code": existing.reference_code
#         })

#     #  Create new signature
#     display_name = (
#         f"{user.first_name} {user.last_name}".strip()
#         or user.email.split("@")[0]
#     )

#     # Generate unique reference code
#     while True:
#         ref_code = ''.join(
#             random.choices(string.ascii_uppercase + string.digits, k=8)
#         )
#         if not SignatureVerification.objects.filter(reference_code=ref_code).exists():
#             break

#     sig = SignatureVerification.objects.create(
#         user=user,
#         display_name=display_name,
#         reference_code=ref_code,
#         ip_address=request.META.get("REMOTE_ADDR"),
#         batch_id=signature_key,  #  UNIQUE PER PART
#         slot=slot,
#         verified_at=timezone.now()
#     )

#     return JsonResponse({
#         "ok": True,
#         "name": sig.display_name,
#         "time": timezone.localtime(sig.verified_at).strftime("%d %b %Y %H:%M"),
#         "code": sig.reference_code
#     })


# =====================================================================================================
# ======================================= Submitted detail ============================================
# =====================================================================================================

@login_required
def submitted_parts_list(request, flow_id):

    batches = FormBatch.objects.filter(flow_id=flow_id)

    if not batches.exists():
        return redirect("submitted_forms_list")

    parts = BatchPart.objects.filter(batch__in=batches,status="ACTIVE")

    if not parts.exists():
        return redirect("submitted_forms_list")

    #  GROUP BY PART ID STRING (NOT PK)
    grouped_parts = (
        FormSubmission.objects
        .filter(part__in=parts )
        .values(
            "part__part_id",                
            "part__batch__batch_id",
            "part__batch__company_name",
        )
        .annotate(
            forms_count=Count("form", distinct=True)
        )
        .order_by("part__part_id")
    )

    formatted_parts = []

    for item in grouped_parts:
        #  Get ONE part pk for linking
        first_part = BatchPart.objects.filter(
            batch__in=batches,
            part_id=item["part__part_id"]
        ).first()

        formatted_parts.append({
            "part_pk": first_part.id,   # single representative
            "part_id": item["part__part_id"],
            "batch_id": item["part__batch__batch_id"],
            "company_name": item["part__batch__company_name"],
            "forms_count": item["forms_count"],
            "created_by": first_part.created_by.email if first_part.created_by else "-",
            "created_at": first_part.created_at,
        })

    return render(
        request,
        "qms_app/form/submitted_parts_list.html",
        {
            "parts": formatted_parts,
            "batch_id": batches.first().batch_id,
            "company_name": batches.first().company_name,
        }
    )




@login_required
def submitted_part_detail(request, part_id):

    part = get_object_or_404(BatchPart, id=part_id)

    rfq_data = {}

    if part.batch.rfq_ref_id:

        rfq_data = build_traceability_data(
            ref_id=part.batch.rfq_ref_id,
            company=part.batch.company_name
        )
    print("WO:", part.batch.batch_id)
    print("RFQ:", part.batch.rfq_ref_id)
    print("RFQ DATA:", rfq_data)
    
    flow_id = part.batch.flow_id
    part_code = part.part_id


    signature_key = f"{flow_id}_{part_code}"

    # Store in session for verify_signature()
    request.session["current_signature_key"] = signature_key

    # --------------------------------------------------
    # 2️⃣ Get ALL batches in same workflow
    # --------------------------------------------------
    batches = FormBatch.objects.filter(flow_id=flow_id)

    # --------------------------------------------------
    # 3️⃣ Get ALL parts with SAME part_id across workflow
    # --------------------------------------------------
    related_parts = BatchPart.objects.filter(
        batch__in=batches,
        part_id=part_code,
        status="ACTIVE"
    )

    # --------------------------------------------------
    # 4️⃣ Get ALL submissions from those parts
    # --------------------------------------------------
    submissions = (
        FormSubmission.objects
        .filter(part__in=related_parts)
        .select_related("stage", "submitted_by", "form")
        .order_by("form__id", "stage__order")
    )

    # --------------------------------------------------
    # 5️⃣ Fetch signatures using UNIQUE PART KEY
    # --------------------------------------------------
    signatures = SignatureVerification.objects.filter(
        batch_id=signature_key
    )

    prepared_sig = signatures.filter(slot="prepared").first()
    approved_sig = signatures.filter(slot="approved").first()

    # --------------------------------------------------
    # If no submissions exist
    # --------------------------------------------------
    if not submissions.exists():
        return render(
            request,
            "qms_app/form/submitted_part_detail.html",
            {
                "part": part,
                "display_forms": {},
                "rfq_data": rfq_data,
                "company_name": part.batch.company_name,
                "submitted_by": None,
                "submitted_at": None,
                "prepared_sig": prepared_sig,
                "approved_sig": approved_sig,
            }
        )

    first_submission = submissions.first()

    # ======================================================
    # GROUP BY FORM → THEN STAGE → THEN FIELDS
    # ======================================================
    display_forms = {}

    for submission in submissions:

        form_name = submission.form.name

        if form_name not in display_forms:
            display_forms[form_name] = []

        stage_fields = []

        for key, value in submission.data.items():

            if not key.startswith("field_"):
                continue

            try:
                field_id = int(key.replace("field_", ""))
                field_obj = FormField.objects.select_related("sub_stage").get(id=field_id)
            except FormField.DoesNotExist:
                continue

            label = field_obj.label
            substage = field_obj.sub_stage.name if field_obj.sub_stage else None
            order = field_obj.order

            # ================= TABLE FIELD =================
            if field_obj.field_type == "table":

                rows = []
                raw_value = value

                if isinstance(raw_value, str):
                    try:
                        raw_value = json.loads(raw_value)
                    except:
                        raw_value = []

                if isinstance(raw_value, list):
                    for row in raw_value:

                        cleaned_cells = []

                        for cell in row.get("cells", []):

                            cell_type = cell.get("type")
                            cell_value = cell.get("value", "")

                            # 🔐 Handle Signature Type
                            if cell_type == "signature" and isinstance(cell_value, dict):

                                cleaned_cells.append({
                                    "type": "signature",
                                    "signed_by": cell_value.get("signed_by"),
                                    "signed_at": cell_value.get("signed_at"),
                                    "reference_code": cell_value.get("reference_code"),
                                })

                            else:
                                if isinstance(cell_value, list):
                                    cell_value = ", ".join(cell_value)

                                cleaned_cells.append({
                                    "type": "normal",
                                    "value": cell_value
                                })
                        rows.append({
                            "name": row.get("name", ""),
                            "cells": cleaned_cells
                        })
                        totals = {}

                        for col in field_obj.table_columns or []:
                            if col.get("is_total"):
                                key = col["name"].lower().replace(" ", "")
                                totals[key] = 0

                        for row in raw_value:
                            for cell in row.get("cells", []):

                                colname = (cell.get("colname") or "").lower()
                                value = cell.get("value")

                                if colname in totals:
                                    try:
                                        totals[colname] += float(value or 0)
                                    except:
                                        pass


                stage_fields.append({
                    "type": "table",
                    "label": label,
                    "columns": field_obj.table_columns or [],
                    "rows": rows,
                    "substage": substage,
                    "order": order,
                    "row_header": getattr(field_obj, "table_row_header", ""),
                    "totals": totals,
                })

            # ================= NORMAL FIELD =================
            else:

                if isinstance(value, list):
                    value = ", ".join(value)

                is_image = field_obj.field_type in ["image_upload", "image_capture"]

                stage_fields.append({
                    "type": "field",
                    "label": label,
                    "value": value,
                    "substage": substage,
                    "order": order,
                    "field_type": field_obj.field_type,
                    "is_image": is_image,
                })

        stage_fields.sort(key=lambda x: x["order"])

        display_forms[form_name].append({
            "stage_name": submission.stage.name,
            "fields": stage_fields,
        })

    # ======================================================

    return render(
        request,
        "qms_app/form/submitted_part_detail.html",
        {
            "part": part,
            "display_forms": display_forms,
            "rfq_data": rfq_data,
            "company_name": part.batch.company_name,
            "lot_number": part.batch.material_batch.lot_number if part.batch.material_batch else "-",
            "submitted_by": first_submission.submitted_by,
            "submitted_at": first_submission.submitted_at,
            "prepared_sig": prepared_sig,
            "approved_sig": approved_sig,
        }
    )


# =============================================================================================================
# =========================================== Delete Details ==================================================
# =============================================================================================================
def delete_form_data(request, form_id):
    form = get_object_or_404(Form, id=form_id)

    if request.method == "POST":
        form.submissions.all().delete()
        return redirect('form_submissions_detail', form_id=form.id)

@login_required
def form_complete(request, form_id):
    form_obj = get_object_or_404(Form, id=form_id)
    return render(request, 'qms_app/form/form_complete.html', {'form_obj': form_obj})

#this view function used to delete the submitted data


@login_required
def delete_submission(request, flow_id):

    if request.method != "POST":
        return redirect("submitted_forms_list")

    #  Get all batches under this workflow
    batches = FormBatch.objects.filter(flow_id=flow_id)

    if not batches.exists():
        return redirect("submitted_forms_list")

    batch_ids = batches.values_list("batch_id", flat=True)

    #  Delete submissions
    FormSubmission.objects.filter(
        submission_batch__in=batch_ids
    ).delete()

    #  Delete signatures
    SignatureVerification.objects.filter(
        batch_id__in=batch_ids
    ).delete()

    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="Form Workflow",
        action="DELETE",
        model_name="Workflow",
        object_repr=str(flow_id),
        ip_address=request.META.get("REMOTE_ADDR")
    )

    #  Delete batches
    batches.delete()

    return redirect("submitted_forms_list")





# Delete Stage
# -----------------------------

def delete_stage(request, stage_id):
    try:
        stage = get_object_or_404(Stage, id=stage_id)

        stage.fields.all().delete()
        stage.sub_stages.all().delete()
        stage.delete()

        return JsonResponse({"status": "success"})

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        })



from django.http import JsonResponse
from .models import Form, Stage

@login_required
def api_form_stages(request, form_id):

    form = Form.objects.get(id=form_id)

    stages = list(
        form.stages.order_by("order")
        .values("id", "name")
    )

    return JsonResponse({
        "form_name": form.name,
        "stages": stages
    })

from django.http import JsonResponse
from .models import Stage, FormField


@login_required
def api_stage_fields(request, stage_id):

    stage = Stage.objects.get(id=stage_id)

    fields = []

    for f in stage.fields.all():

        fields.append({
            "id": f.id,
            "label": f.label,
            "type": f.field_type,
            "options": f.options,
            "table_columns": f.table_columns,
            "table_rows": f.table_rows,
            "required": f.required,
        })


    return JsonResponse({
        "stage": stage.name,
        "fields": fields
    })


@login_required
def flow_kanban(request):

    forms = Form.objects.all()

    return render(request,"qms_app/flow/kanban.html",{
        "forms":forms
    })



# ===============================================================================================
# ========================================= Settings ============================================
# ===============================================================================================




@login_required
def page_settings(request):

    User = get_user_model()
    users = User.objects.all().select_related("page_access")
    all_folders = FormFolder.objects.all()

    if request.method == "POST":

        for user in users:

            #  Skip superusers (always full access)
            if user.is_superuser:
                continue

            # Get or create access object
            access, created = UserPageAccess.objects.get_or_create(user=user)

            # ================= CORE PAGE ACCESS =================
            access.can_dashboard = f"dashboard_{user.id}" in request.POST
            access.can_forms = f"forms_{user.id}" in request.POST
            access.can_submitted = f"submitted_{user.id}" in request.POST
            access.can_documents = f"documents_{user.id}" in request.POST
            access.can_sop = f"sop_{user.id}" in request.POST
            access.can_userdetail = f"userdetail_{user.id}" in request.POST
            access.can_materialBatch = f"materialBatch_{user.id}" in request.POST
            access.can_form_build = f"form_build_{user.id}" in request.POST
            access.can_fm = f"fm_{user.id}" in request.POST
            access.can_costing_dashboard = f"costing_dash_{user.id}" in request.POST
            access.can_workflow_access = f"dworkflow_{user.id}" in request.POST
            # ================= MODULE ACCESS =================
            access.can_goods_entry = f"goods_{user.id}" in request.POST
            access.can_tests = f"tests_{user.id}" in request.POST

            access.can_add = f"add_{user.id}" in request.POST
            access.can_edit = f"edit_{user.id}" in request.POST
            access.can_delete = f"delete_{user.id}" in request.POST
            access.can_signature_btn = f"sbtn_{user.id}" in request.POST
            access.can_approved_document = f"ad_{user.id}" in request.POST
            
            access.can_sign_prepared = f"sign_prepared_{user.id}" in request.POST
            access.can_sign_reviewed = f"sign_reviewed_{user.id}" in request.POST
            access.can_sign_approved = f"sign_approved_{user.id}" in request.POST
            
            access.save()

            # ================= FOLDER ACCESS =================
            folder_ids = request.POST.getlist(f"folders_{user.id}")
            access.allowed_folders.set(folder_ids)

        return redirect("page_settings")

    return render(
        request,
        "qms_app/settings/page_settings.html",
        {
            "users": users,
            "all_folders": all_folders,
        }
    )



@login_required
def unlock_user(request, user_id):

    if not request.user.is_superuser:
        return JsonResponse({"status": "error"})

    User = get_user_model()
    user = User.objects.get(id=user_id)

    user.failed_attempts = 0
    user.account_locked = False
    user.save()

    return JsonResponse({"status": "success"})


# ============================================================================================================
# ============================================= CAPA =========================================================
# ============================================================================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.utils import timezone
from .models import CAPA
from .forms import CAPAForm
from audit_log.models import AuditLog


# =====================================================
# CAPA LIST VIEW (Dashboard Ready)
# =====================================================
@login_required
def capa_list(request):

    # Role-Based Visibility
    if request.user.role == "admin":
        capas = CAPA.objects.all()
    else:
        capas = CAPA.objects.filter(
            Q(created_by=request.user) |
            Q(assigned_to=request.user)
        )

    capas = capas.select_related(
        "assigned_to",
        "related_batch_part",
        "related_ncr"
    ).order_by("-created_at")

    # KPI Counters (For Dashboard)
    total = capas.count()
    open_count = capas.filter(status="OPEN").count()
    overdue_count = sum(1 for c in capas if c.is_overdue())
    critical_count = capas.filter(priority="CRITICAL").count()

    return render(
        request,
        "qms_app/capa/capa_list.html",
        {
            "capas": capas,
            "capa_statuses": CAPA.STATUS_CHOICES,
            "total": total,
            "open_count": open_count,
            "overdue_count": overdue_count,
            "critical_count": critical_count,
        }
    )


# =====================================================
# SAFE STATUS UPDATE (AJAX)
# =====================================================
@require_POST
@login_required
def update_capa_status(request):

    capa_id = request.POST.get("capa_id")
    new_status = request.POST.get("status")

    capa = get_object_or_404(CAPA, id=capa_id)

    # Permission check
    if request.user.role != "admin" and request.user != capa.assigned_to:
        return HttpResponseForbidden("You do not have permission.")

    if new_status not in dict(CAPA.STATUS_CHOICES):
        return JsonResponse({"success": False, "error": "Invalid status"})

    old_status = capa.status
    capa.status = new_status
    capa.save()

    # Audit Log
    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="CAPA",
        action="STATUS_UPDATE",
        model_name="CAPA",
        object_repr=capa.capa_number,
        old_data={"status": old_status},
        new_data={"status": new_status},
        ip_address=request.META.get("REMOTE_ADDR")
    )

    return JsonResponse({"success": True})


# =====================================================
# CREATE CAPA
# =====================================================
@login_required
def capa_create(request):

    if request.method == "POST":
        form = CAPAForm(request.POST)

        if form.is_valid():
            capa = form.save(commit=False)
            capa.created_by = request.user
            capa.status = "OPEN"
            capa.save()

            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="CAPA",
                action="CREATE",
                model_name="CAPA",
                object_repr=capa.capa_number,
                new_data={"title": capa.title},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect("capa_detail", pk=capa.pk)

    else:
        form = CAPAForm()

    return render(
        request,
        "qms_app/capa/capa_create.html",
        {"form": form}
    )


# =====================================================
# CAPA DETAIL VIEW
# =====================================================
@login_required
def capa_detail(request, pk):

    capa = get_object_or_404(CAPA, pk=pk)

    # Permission check
    if request.user.role != "admin" and \
       request.user != capa.assigned_to and \
       request.user != capa.created_by:
        return HttpResponseForbidden("Access Denied")

    if request.method == "POST":

        form = CAPAForm(request.POST, instance=capa)

        if form.is_valid():
            old_status = capa.status
            updated_capa = form.save()

            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="CAPA",
                action="UPDATE",
                model_name="CAPA",
                object_repr=updated_capa.capa_number,
                old_data={"status": old_status},
                new_data={"status": updated_capa.status},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect("capa_detail", pk=capa.pk)

    else:
        form = CAPAForm(instance=capa)

    return render(
        request,
        "qms_app/capa/capa_detail.html",
        {
            "capa": capa,
            "form": form,
            "is_overdue": capa.is_overdue(),
        }
    )


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.utils import timezone

from .models import CAPA
from .forms import CAPAForm
from audit_log.models import AuditLog


# =====================================================
# CAPA LIST VIEW (Dashboard Ready)
# =====================================================
@login_required
def capa_list(request):

    # Role-Based Visibility
    if request.user.role == "admin":
        capas = CAPA.objects.all()
    else:
        capas = CAPA.objects.filter(
            Q(created_by=request.user) |
            Q(assigned_to=request.user)
        )

    capas = capas.select_related(
        "assigned_to",
        "related_batch_part",
        "related_ncr"
    ).order_by("-created_at")

    # KPI Counters (For Dashboard)
    total = capas.count()
    open_count = capas.filter(status="OPEN").count()
    overdue_count = sum(1 for c in capas if c.is_overdue())
    critical_count = capas.filter(priority="CRITICAL").count()

    return render(
        request,
        "qms_app/capa/capa_list.html",
        {
            "capas": capas,
            "capa_statuses": CAPA.STATUS_CHOICES,
            "total": total,
            "open_count": open_count,
            "overdue_count": overdue_count,
            "critical_count": critical_count,
        }
    )


# =====================================================
# SAFE STATUS UPDATE (AJAX)
# =====================================================
@require_POST
@login_required
def update_capa_status(request):

    capa_id = request.POST.get("capa_id")
    new_status = request.POST.get("status")

    capa = get_object_or_404(CAPA, id=capa_id)

    # Permission check
    if request.user.role != "admin" and request.user != capa.assigned_to:
        return HttpResponseForbidden("You do not have permission.")

    if new_status not in dict(CAPA.STATUS_CHOICES):
        return JsonResponse({"success": False, "error": "Invalid status"})

    old_status = capa.status
    capa.status = new_status
    capa.save()

    # Audit Log
    AuditLog.objects.create(
        user=request.user,
        role=request.user.role,
        module="CAPA",
        action="STATUS_UPDATE",
        model_name="CAPA",
        object_repr=capa.capa_number,
        old_data={"status": old_status},
        new_data={"status": new_status},
        ip_address=request.META.get("REMOTE_ADDR")
    )

    return JsonResponse({"success": True})


# =====================================================
# CREATE CAPA
# =====================================================
@login_required
def capa_create(request):

    if request.method == "POST":
        form = CAPAForm(request.POST)

        if form.is_valid():
            capa = form.save(commit=False)
            capa.created_by = request.user
            capa.status = "OPEN"
            capa.save()

            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="CAPA",
                action="CREATE",
                model_name="CAPA",
                object_repr=capa.capa_number,
                new_data={"title": capa.title},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect("capa_detail", pk=capa.pk)

    else:
        form = CAPAForm()

    return render(
        request,
        "qms_app/capa/capa_create.html",
        {"form": form}
    )


# =====================================================
# CAPA DETAIL VIEW
# =====================================================
@login_required
def capa_detail(request, pk):

    capa = get_object_or_404(CAPA, pk=pk)

    # Permission check
    if request.user.role != "admin" and \
       request.user != capa.assigned_to and \
       request.user != capa.created_by:
        return HttpResponseForbidden("Access Denied")

    if request.method == "POST":

        form = CAPAForm(request.POST, instance=capa)

        if form.is_valid():
            old_status = capa.status
            updated_capa = form.save()

            AuditLog.objects.create(
                user=request.user,
                role=request.user.role,
                module="CAPA",
                action="UPDATE",
                model_name="CAPA",
                object_repr=updated_capa.capa_number,
                old_data={"status": old_status},
                new_data={"status": updated_capa.status},
                ip_address=request.META.get("REMOTE_ADDR")
            )

            return redirect("capa_detail", pk=capa.pk)

    else:
        form = CAPAForm(instance=capa)

    return render(
        request,
        "qms_app/capa/capa_detail.html",
        {
            "capa": capa,
            "form": form,
            "is_overdue": capa.is_overdue(),
        }
    )


# ==================================================NCR====================================================


@require_POST
@login_required
def move_to_ncr(request):

    try:
        data = json.loads(request.body)
        part_id = data.get("part_id")

        if not part_id:
            return JsonResponse({
                "success": False,
                "error": "Missing part_id"
            })

        part = BatchPart.objects.select_related("batch", "current_stage").get(id=part_id)

        # Prevent duplicate NCR
        if part.status == "NCR":
            return JsonResponse({
                "success": False,
                "error": "Part already in NCR"
            })

        #  1. Update Part Status
        part.status = "NCR"
        part.save(update_fields=["status"])

        #  2. Create NCR Record
        ncr = NCR.objects.create(
            batch_part=part,
            form=part.batch.form,
            stage=part.current_stage,
            description=f"Auto-created from Kanban drag by {request.user.email}",
            severity=1,
            occurrence=1,
            detection=1,
            created_by=request.user
        )

        #  3. Audit Log
        AuditLog.objects.create(
            user=request.user,
            role=request.user.role,
            module="NCR",
            action="AUTO_CREATE",
            model_name="NCR",
            object_repr=ncr.ncr_number,
            new_data={
                "batch_id": part.batch.batch_id,
                "part_id": part.part_id,
                "stage": part.current_stage.name
            },
            ip_address=request.META.get("REMOTE_ADDR")
        )

        return JsonResponse({"success": True})

    except BatchPart.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Part not found"
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        })
    


@login_required
def create_capa_from_ncr(request, ncr_id):

    ncr = get_object_or_404(NCR, id=ncr_id)

    # Prevent duplicate CAPA
    existing = CAPA.objects.filter(related_ncr=ncr).first()
    if existing:
        return redirect("capa_detail", pk=existing.id)

    # Update NCR status
    ncr.status = "CAPA_REQUIRED"
    ncr.save(update_fields=["status"])

    # Create CAPA
    capa = CAPA.objects.create(
        related_ncr=ncr,
        related_batch_part=ncr.batch_part,
        title=f"CAPA for {ncr.ncr_number}",
        problem_statement=ncr.description,
        created_by=request.user,
        assigned_to=request.user
    )

    return redirect("capa_detail", pk=capa.id)



from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

@require_POST
@login_required
def delete_ncr(request, ncr_id):

    ncr = get_object_or_404(NCR, id=ncr_id)

    part = ncr.batch_part

    #  Restore part status
    part.status = "ACTIVE"
    part.save(update_fields=["status"])

    # 🗑 Delete NCR record
    ncr.delete()

    return JsonResponse({"success": True})






@require_page_permission("can_materialBatch")
@login_required
def material_batches(request):

    if request.method == "POST":

        material_name = request.POST.get("material_name")
        supplier = request.POST.get("supplier")
        lot_number = request.POST.get("lot_number")
        received_date = request.POST.get("received_date")

        if not material_name or not lot_number:
            messages.error(request, "Material name and Lot number required.")
            return redirect("material_batches")

        MaterialBatch.objects.create(
            material_name=material_name,
            supplier=supplier,
            lot_number=lot_number,
            received_date=received_date,
            created_by=request.user
        )

        messages.success(request, "Material batch created.")

        return redirect("material_batches")

    batches = MaterialBatch.objects.select_related(
        "created_by"
    ).order_by("-received_date")

    return render(
        request,
        "qms_app/material/material_batches.html",
        {
            "batches": batches
        }
    )



# ========================================================================================
# ================================ User Work Trace ======================================= 
# ========================================================================================
@login_required
@require_page_permission("can_userdetail")
def user_work_trace(request):

    import json

    # ==============================================
    # GET ALL BATCHES ORDERED
    # ==============================================

    all_batches = FormBatch.objects.select_related(
        "material_batch",
        "created_by"
    ).order_by("-created_at")

    # ==============================================
    # REMOVE DUPLICATE FLOW_ID
    # ==============================================

    seen_flows = set()
    batches = []

    for batch in all_batches:
        if batch.flow_id not in seen_flows:
            batches.append(batch)
            seen_flows.add(batch.flow_id)

    # ==============================================
    # COMPANY DROPDOWN
    # ==============================================

    companies = (
        FormBatch.objects
        .values_list("company_name", flat=True)
        .distinct()
    )

    # ==============================================
    # BUILD JSON FOR UI
    # ==============================================

    batches_json = []

    for batch in batches:

        parts_queryset = BatchPart.objects.filter(
            batch__flow_id=batch.flow_id
        ).select_related(
            "batch",
            "batch__material_batch",
            "created_by"
        ).order_by("part_id", "-created_at")

        # ------------------------------------------
        # KEEP ONLY LATEST PART PER PART_ID
        # ------------------------------------------

        latest_parts = {}

        for p in parts_queryset:
            if p.part_id not in latest_parts:
                latest_parts[p.part_id] = p

        parts_data = []
        history_data = []

        for part in latest_parts.values():

            lot_number = "-"

            if part.batch.material_batch:
                lot_number = part.batch.material_batch.lot_number

            parts_data.append({
                "id": part.id,
                "part_id": part.part_id,
                "lot": lot_number,
                "user": part.created_by.email if part.created_by else "-",
                "status": part.get_status_display(),
                "created": timezone.localtime(part.created_at).strftime("%d %b %Y %H:%M")
            })

            # ------------------------------------------
            # STAGE HISTORY PER PART
            # ------------------------------------------

            histories = StageHistory.objects.filter(
                part=part
            ).select_related(
                "from_stage",
                "to_stage",
                "to_stage__form",
                "moved_by"
            ).order_by("moved_at")

            for h in histories:

                history_data.append({
                    "form": h.to_stage.form.name if h.to_stage and h.to_stage.form else "-",
                    "from_stage": h.from_stage.name if h.from_stage else "-",
                    "to_stage": h.to_stage.name if h.to_stage else "-",
                    "user": h.moved_by.email if h.moved_by else "-",
                    "time": timezone.localtime(h.moved_at).strftime("%d %b %Y %H:%M")
                })

        batches_json.append({
            "id": batch.id,
            "company": batch.company_name,
            "batch_id": batch.batch_id,
            "parts": parts_data,
            "history": history_data
        })

    return render(
        request,
        "qms_app/user_work_trace.html",
        {
            "batches": batches,
            "companies": companies,
            "batches_json": json.dumps(batches_json)
        }
    )
    
# ============================================================================================================================
# =====================================================Part Audit Log ========================================================
# ============================================================================================================================
# @login_required
# def part_audit_logs(request, part_id):

#     part = get_object_or_404(BatchPart, id=part_id)

#     logs = AuditLog.objects.filter(
#         model_name="BatchPart",
#         object_id=str(part.id)
#     ).select_related("user").order_by("timestamp")

#     data = []

#     for log in logs:

#         old_data = log.old_data or {}
#         new_data = log.new_data or {}

#         form_name = new_data.get("form") or old_data.get("form") or "-"
#         stage_name = new_data.get("stage") or old_data.get("stage") or "-"

#         # Fix old logs where IDs were stored instead of names
#         if isinstance(form_name, int) or str(form_name).isdigit():
#             try:
#                 form_name = Form.objects.get(id=form_name).name
#             except:
#                 pass

#         if isinstance(stage_name, int) or str(stage_name).isdigit():
#             try:
#                 stage_name = Stage.objects.get(id=stage_name).name
#             except:
#                 pass

#         data.append({
#             "action": log.action,
#             "user": log.user.email if log.user else "-",
#             "form": form_name,
#             "stage": stage_name,
#             "time": timezone.localtime(log.timestamp).strftime("%d %b %Y %H:%M")
#         })

#     return JsonResponse({"logs": data})

@login_required
def part_audit_logs(request, part_id):

    part = get_object_or_404(BatchPart, id=part_id)

    logs = (
        AuditLog.objects
        .filter(
            object_id=str(part.id),
            action__in=[
                "CREATE",
                "STAGE_SUBMITTED",
                "STAGE_MOVED",
                "FORM_MOVED"
            ]
        )
        .exclude(action="CREATE", module="Form Submission")
        .select_related("user")
        .order_by("timestamp")
    )

    data = []

    for log in logs:

        old_data = log.old_data or {}
        new_data = log.new_data or {}

        old_form = old_data.get("form")
        new_form = new_data.get("form")

        old_stage = old_data.get("stage")
        new_stage = new_data.get("stage")

        folder_name = "-"

        # ------------------------------------------
        # Resolve folder from form name
        # ------------------------------------------

        try:
            form_lookup = new_form or old_form

            if form_lookup:

                # if stored as ID
                if isinstance(form_lookup, int) or str(form_lookup).isdigit():
                    form_obj = Form.objects.select_related("folder").filter(id=form_lookup).first()

                # if stored as NAME
                else:
                    form_obj = Form.objects.select_related("folder").filter(name=form_lookup).first()

                if form_obj and form_obj.folder:
                    folder_name = form_obj.folder.name

        except:
            pass

        # ------------------------------------------
        # Resolve STAGE names
        # ------------------------------------------

        if isinstance(old_stage, int) or str(old_stage).isdigit():
            try:
                old_stage = Stage.objects.get(id=old_stage).name
            except:
                pass

        if isinstance(new_stage, int) or str(new_stage).isdigit():
            try:
                new_stage = Stage.objects.get(id=new_stage).name
            except:
                pass

        # ------------------------------------------
        # Build FORM display
        # ------------------------------------------

        if old_form and new_form and old_form != new_form:
            form_name = f"{old_form} → {new_form}"
        else:
            form_name = new_form or old_form or "-"

        # ------------------------------------------
        # Build STAGE display
        # ------------------------------------------

        if log.action == "STAGE_MOVED" and old_stage and new_stage and old_stage != new_stage:
            stage_name = f"{old_stage} → {new_stage}"
        else:
            stage_name = new_stage or old_stage or "-"

        # ------------------------------------------
        # Skip bad logs
        # ------------------------------------------

        if str(form_name).isdigit() or str(stage_name).isdigit():
            continue

        # ------------------------------------------
        # Append timeline record
        # ------------------------------------------

        data.append({
            "action": log.action,
            "user": log.user.email if log.user else "-",
            "folder": folder_name,
            "form": form_name,
            "stage": stage_name,
            "time": timezone.localtime(log.timestamp).strftime("%d %b %Y %H:%M")
        })

    return JsonResponse({"logs": data})




from .models import MachineSession
from django.utils import timezone
import json


@login_required
def machine_start(request):

    data = json.loads(request.body)

    password = data.get("password")
    machine_id = data.get("machine_id")
    message = data.get("message")

    if not request.user.check_password(password):
        return JsonResponse({"success": False})

    shift = "Morning"

    hour = timezone.localtime().hour

    if hour >= 14 and hour < 22:
        shift = "Evening"
    elif hour >= 22 or hour < 6:
        shift = "Night"

    MachineSession.objects.create(
        machine_id=machine_id,
        operator=request.user,
        shift=shift,
        message=message,
        ip_address=request.META.get("REMOTE_ADDR")
    )

    request.session["machine_verified"] = True

    return JsonResponse({"success": True})

import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required


from django.contrib.auth import authenticate
import json

@login_required
def machine_verify(request):

    data = json.loads(request.body)

    password = data.get("password")

    user = authenticate(
        request,
        email=request.user.email,
        password=password
    )

    if user:
        return JsonResponse({"success": True})

    return JsonResponse({"success": False})

from django.http import JsonResponse
from django.utils.timezone import localtime
from .models import MachineSession


def machine_previous(request):

    machine_id = request.GET.get("machine_id")

    if not machine_id:
        return JsonResponse({})

    session = (
        MachineSession.objects
        .filter(machine_id__iexact=machine_id)
        .exclude(message__isnull=True)
        .exclude(message__exact="")
        .exclude(message="-")
        .order_by("-logout_time")
        .first()
    )

    if not session:
        return JsonResponse({})

    return JsonResponse({
        "previous_user": session.operator.email,
        "shift": session.shift,
        "login_time": session.login_time.strftime("%d %b %H:%M"),
        "message": session.message
    })
    
from django.utils import timezone

@login_required
def machine_logout(request):

    data = json.loads(request.body)

    message = data.get("message")

    session = MachineSession.objects.filter(
        operator=request.user,
        logout_time__isnull=True
    ).last()

    if session:

        session.logout_time = timezone.now()

        if message:
            session.message = message

        session.save()

    return JsonResponse({"success": True})

@login_required
def machine_check_session(request):

    active = MachineSession.objects.filter(
        operator=request.user,
        logout_time__isnull=True
    ).exists()

    return JsonResponse({"active": active})


from django.http import JsonResponse
from django.utils.timezone import localtime
from .models import MachineSession
def machine_log(request):

    machine_id = request.GET.get("machine_id")

    if not machine_id:
        return JsonResponse({"logs": []})

    sessions = (
        MachineSession.objects
        .filter(machine_id__iexact=machine_id)
        .exclude(message__isnull=True)
        .exclude(message__exact="")
        .exclude(message="-")
        .order_by("-logout_time")[:5]
    )

    logs = []

    for s in sessions:
        logs.append({
            "operator": s.operator.email,
            "time": localtime(s.logout_time).strftime("%H:%M"),
            "message": s.message
        })

    return JsonResponse({"logs": logs})




def machine_dashboard(request):

    machines = MachineSession.objects.values_list(
        "machine_id", flat=True
    ).distinct()

    selected_machine = request.GET.get("machine")

    logs = []

    if selected_machine:
        logs = MachineSession.objects.filter(
            machine_id=selected_machine
        ).order_by("-logout_time")

    active_sessions = MachineSession.objects.filter(
        logout_time__isnull=True
    )

    return render(request, "qms_app/machine_dashboard.html", {
        "machines": machines,
        "logs": logs,
        "selected_machine": selected_machine,
        "active_sessions": active_sessions
    })





# ==========================================================
# ======================================================
# =========================================================


from form_builder.models import FormResponse

@login_required
def rfq_to_workorder_list(request):

    rfqs = (
        FormResponse.objects
        .filter(
            ref_id__startswith="RFQ",
            status="WON"
        )
        .exclude(company__isnull=True)
        .exclude(company="")
        .select_related("form")
        .order_by("-created_at")
    )

    return render(
        request,
        "qms_app/work_order/rfq_to_workorder_list.html",
        {
            "rfqs": rfqs
        }
    )
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render

from form_builder.models import FormResponse
from po_qu.models import WorkOrder


User = get_user_model()


# ==========================================================
# RFQ → WORKORDER LIST
# ==========================================================

@login_required
def rfq_to_workorder_list(request):

    rfqs = (
        FormResponse.objects
        .filter(
            ref_id__startswith="RFQ",
            status__in=["WON", "WORKORDER_CREATED"]
        )
        .exclude(company__isnull=True)
        .exclude(company="")
        .select_related("form")
        .order_by("-created_at")
    )

    converted_rfqs = WorkOrder.objects.values_list(
        "rfq_ref",
        flat=True
    )

    return render(
        request,
        "qms_app/work_order/rfq_to_workorder_list.html",
        {
            "rfqs": rfqs,
            "converted_rfqs": converted_rfqs
        }
    )

# ==========================================================
# WORKORDER LIST
# ==========================================================

# @login_required
# def workorder_list(request):

#     workorders = (
#         WorkOrder.objects
#         .all()
#         .order_by("-created_at")
#     )

#     return render(
#         request,
#         "qms_app/work_order/workorder_list.html",
#         {
#             "workorders": workorders
#         }
#     )
@login_required
def workorder_list(request):

    # =====================================================
    # ONLY APPROVED WORKORDERS
    # =====================================================

    workorders = (

        WorkOrder.objects

        .filter(
            approved_by__isnull=False
        )

        .prefetch_related("items")

        .order_by("-id")

    )

    # =====================================================
    # STATS
    # =====================================================

    total_workorders = workorders.count()

    total_parts = sum(
        wo.items.count()
        for wo in workorders
    )

    total_customers = len(

        set(
            wo.customer_name
            for wo in workorders
        )

    )

    # =====================================================
    # PAGE
    # =====================================================

    return render(

        request,

        "qms_app/work_order/workorder_list.html",

        {

            "workorders": workorders,

            "total_workorders": total_workorders,

            "total_parts": total_parts,

            "total_customers": total_customers,

        }

    )





from po_qu.models import (WorkOrder,WorkOrderItem)
from qms_app.models import (Form,WorkOrderPartFormSubmission,FormBatch,BatchPart)
from po_qu.models import RFQ
from django.shortcuts import (render,get_object_or_404)
from django.contrib.auth.decorators import (login_required)
from qms_app.models import FormSubmission
from django.contrib.auth import get_user_model
User = get_user_model()

@login_required
def workorder_kanban(request, wo_id):

    # =====================================================
    # GET WORK ORDER
    # =====================================================

    wo = get_object_or_404(

        WorkOrder,

        id=wo_id

    )

    # =====================================================
    # GET ITEMS
    # =====================================================

    items = (

        wo.items

        .all()

        .select_related(

            "assigned_to",

            "stage_folder"

        )

    )

    # =====================================================
    # STAGE COLUMNS
    # =====================================================

    stages = {

        "INTERNAL_WORKORDER": [],

        "RFQ_PREP": [],

        "RAW_MATERIAL_PO": [],

        "RAW_MATERIAL_RECEIVED": [],

        "PRODUCTION": [],

        "FINAL_INSPECTION": [],

        "DISPATCH": [],

        "POST_DISPATCH": [],

    }

    # =====================================================
    # PROCESS ITEMS
    # =====================================================

    for item in items:
        # =================================================
        # RFQ DETAILS
        # =================================================

        item.rfq = None

        if item.rfq_id:

            try:

                item.rfq = RFQ.objects.prefetch_related(
                    "attachments"
                ).get(

                    rfq_number=item.rfq_id
                )

            except RFQ.DoesNotExist:

                pass
        # =================================================
        # GET FORMS
        # =================================================

        forms = Form.objects.filter(
            folder=item.stage_folder
        ).distinct()

        # =================================================
        # GET COMPLETED FORM IDS
        # =================================================

        completed_form_ids = []

        for form_obj in forms:

            # =========================================
            # TOTAL STAGES IN FORM
            # =========================================

            total_stages = (
                form_obj.stages.count()
            )

            # =========================================
            # COMPLETED STAGES
            # =========================================

            completed_stages = (

                FormSubmission.objects.filter(

                    form=form_obj,

                    part__part_id=item.product_code

                )

                .values("stage")

                .distinct()

                .count()

            )
            # =========================================
            # STORE COUNTS FOR TEMPLATE
            # =========================================

            form_obj.total_stages = (
                total_stages
            )

            form_obj.completed_stages = (
                completed_stages
            )
            # =========================================
            # FULLY COMPLETED
            # =========================================

            if (

                total_stages > 0

                and

                completed_stages >= total_stages

            ):

                completed_form_ids.append(
                    form_obj.id
                )
        # =================================================
        # COUNTS
        # =================================================

        completed_forms = len(
            completed_form_ids
        )

        required_forms = forms.count()

        pending_forms = (
            required_forms - completed_forms
        )

        # =================================================
        # CAN MOVE
        # =================================================

        can_move = (

            required_forms > 0

            and

            completed_forms >= required_forms

        )

        # =================================================
        # TEMPLATE VALUES
        # =================================================

        item.forms_list = forms

        item.completed_form_ids = (
            completed_form_ids
        )

        item.completed_forms = (
            completed_forms
        )

        item.required_forms = (
            required_forms
        )

        item.pending_forms = (
            pending_forms
        )

        item.can_move = (
            can_move
        )

        # =================================================
        # ADD TO STAGE COLUMN
        # =================================================

        stages[item.status].append(item)

    # =====================================================
    # RENDER PAGE
    # =====================================================

    return render(

        request,

        "qms_app/work_order/workorder_kanban.html",

        {

            "wo": wo,

            "stages": stages,
            "users": User.objects.all(),

        }

    )




# ================================================================================
# RFQ ATTACHMENT UPLOAD( if we get any quotation for that rfq it will added here )
# =========================================================

@login_required
def upload_rfq_attachment(request, rfq_id):

    rfq = get_object_or_404(

        RFQ,

        id=rfq_id
    )

    if request.method == "POST":

        uploaded_file = request.FILES.get(
            "attachment"
        )

        if uploaded_file:

            RFQAttachment.objects.create(

                rfq=rfq,

                file=uploaded_file,

                uploaded_by=request.user

            )

            messages.success(

                request,

                "Attachment Uploaded Successfully"
            )

        else:

            messages.error(

                request,

                "No File Selected"
            )

    return redirect(

        request.META.get(
            "HTTP_REFERER",
            "/"
        )

    )


@login_required
def create_rfq_from_workorder(request, item_id):

    item = get_object_or_404(
        WorkOrderItem,
        id=item_id
    )

    # STORE ITEM ID IN SESSION

    request.session["rfq_workorder_item_id"] = item.id

    # REDIRECT TO REAL RFQ CREATE PAGE

    return redirect("create_rfq")


from po_qu.models import RFQ


@login_required
def rfq_detail(request, rfq_id):

    rfq = get_object_or_404(
        RFQ,
        id=rfq_id
    )

    return render(
        request,
        "qms_app/work_order/rfq_detail.html",
        {
            "rfq": rfq
        }
    )

@login_required
def assign_workorder_user(request):

    if request.method != "POST":

        return JsonResponse({

            "status": "error"

        })

    data = json.loads(request.body)

    item_id = data.get("item_id")

    user_id = data.get("user_id")

    item = get_object_or_404(

        WorkOrderItem,

        id=item_id

    )

    # ============================================
    # ASSIGN USER
    # ============================================

    if user_id:

        item.assigned_to = (
            User.objects.get(id=user_id)
        )

    else:

        item.assigned_to = None

    item.save()

    return JsonResponse({

        "status": "success"

    })


@login_required
def stage_forms(request, item_id):

    item = get_object_or_404(
        WorkOrderItem,
        id=item_id
    )

    forms = Form.objects.filter(
        folder=item.stage_folder
    ).distinct()

    completed_form_ids = (
        WorkOrderPartFormSubmission.objects.filter(
            workorder_item=item
        ).values_list(
            "form_id",
            flat=True
        )
    )

    return render(

        request,

        "qms_app/work_order/stage_forms.html",

        {

            "item": item,

            "forms": forms,

            "completed_form_ids": completed_form_ids,

        }

    )


import json

from django.http import JsonResponse

from qms_app.models import FormFolder


@login_required
def update_workorder_stage(request):

    if request.method == "POST":

        try:

            # =====================================
            # LOAD JSON DATA
            # =====================================

            data = json.loads(
                request.body
            )

            item_id = data.get(
                "item_id"
            )

            new_stage = data.get(
                "stage"
            )

            print(
                "ITEM ID:",
                item_id
            )

            print(
                "NEW STAGE:",
                new_stage
            )

            # =====================================
            # GET ITEM
            # =====================================

            item = get_object_or_404(

                WorkOrderItem,

                id=item_id

            )

            # =====================================
            # VALID STAGES
            # =====================================

            valid_stages = [

                "INTERNAL_WORKORDER",

                "RFQ_PREP",

                "RAW_MATERIAL_PO",

                "RAW_MATERIAL_RECEIVED",

                "PRODUCTION",

                "FINAL_INSPECTION",

                "DISPATCH",

                "POST_DISPATCH",

            ]

            # =====================================
            # INVALID STAGE
            # =====================================

            if new_stage not in valid_stages:

                return JsonResponse({

                    "success": False,

                    "error": "Invalid Stage"

                })

            # =====================================
            # UPDATE STATUS
            # =====================================

            item.status = new_stage

            # =====================================
            # MAP STAGE → FOLDER
            # =====================================

            folder_mapping = {

                "INTERNAL_WORKORDER":
                    "Internal Workorder",

                "RFQ_PREP":
                    "RFQ Preparation",

                "RAW_MATERIAL_PO":
                    "Raw Material Purchase Order",

                "RAW_MATERIAL_RECEIVED":
                    "Receive Raw material",

                "PRODUCTION":
                    "Production",

                "FINAL_INSPECTION":
                    "Final Inspection",

                "DISPATCH":
                    "Dispatch",

                "POST_DISPATCH":
                    "Post dispatch & Feed back",

            }

            # =====================================
            # GET FOLDER NAME
            # =====================================

            folder_name = folder_mapping.get(
                new_stage
            )

            print(
                "FOLDER NAME:",
                folder_name
            )

            # =====================================
            # GET FOLDER
            # =====================================

            folder = FormFolder.objects.filter(

                name__iexact=folder_name

            ).first()

            print(
                "FOUND FOLDER:",
                folder
            )

            # =====================================
            # UPDATE STAGE FOLDER
            # =====================================

            item.status = new_stage
            item.stage_folder = folder
            item.save()

            item.refresh_from_db()

            print(
                "UPDATED STATUS:",
                item.status
            )

            return JsonResponse({

                "success": True,

                "stage": item.status,

                "folder":
                    folder.name if folder else None

            })

        except Exception as e:

            print(
                "STAGE UPDATE ERROR:",
                str(e)
            )

            return JsonResponse({

                "success": False,

                "error": str(e)

            })

    return JsonResponse({

        "success": False,

        "error": "Invalid Request"

    })






import uuid

from django.shortcuts import (
    redirect,
    get_object_or_404
)

from django.contrib.auth.decorators import (
    login_required
)

from po_qu.models import (
    WorkOrderItem
)

from qms_app.models import (

    Form,

    FormBatch,

    BatchPart

)



# @login_required
# def start_workorder_form(

#     request,

#     form_id,

#     item_id

# ):

#     # =====================================
#     # GET FORM
#     # =====================================

#     form_obj = get_object_or_404(

#         Form,

#         id=form_id

#     )

#     # =====================================
#     # GET WORKORDER ITEM
#     # =====================================

#     item = get_object_or_404(

#         WorkOrderItem,

#         id=item_id

#     )

#     # =====================================
#     # GET FIRST STAGE
#     # =====================================

#     first_stage = (

#         form_obj.stages

#         .order_by("order")

#         .first()

#     )

#     # =====================================
#     # NO STAGE
#     # =====================================

#     if not first_stage:

#         return redirect(
   
#             "workorder_kanban",

#             wo_id=item.work_order.id

#         )

#     # =====================================
#     # PART DISPLAY
#     # =====================================

#     part_display = item.product_code

#     # =====================================
#     # CREATE / GET FORM BATCH
#     # =====================================

#     form_batch, created = (

#         FormBatch.objects.get_or_create(

#             form=form_obj,

#             batch_id=item.work_order.wo_number,

#             defaults={

#                 "company_name":
#                     item.work_order.customer_name,

#                 "flow_id":
#                     uuid.uuid4(),

#                 "created_by":
#                     request.user,

#                 "current_stage":
#                     first_stage

#             }

#         )

#     )

#     # =====================================
#     # ALWAYS UPDATE COMPANY NAME
#     # =====================================

#     form_batch.company_name = (
#         item.work_order.customer_name
#     )

#     form_batch.current_stage = (
#         first_stage
#     )

#     form_batch.save()

#     # =====================================
#     # GET EXISTING BATCH PART
#     # =====================================

#     batch_part = BatchPart.objects.filter(

#         batch=form_batch,

#         part_id=part_display

#     ).first()

#     # =====================================
#     # CREATE IF NOT EXISTS
#     # =====================================

#     if not batch_part:

#         batch_part = BatchPart.objects.create(

#             batch=form_batch,

#             part_id=part_display,

#             current_stage=first_stage

#         )

#     # =====================================
#     # UPDATE IF EXISTS
#     # =====================================

#     else:

#         batch_part.batch = form_batch

#         batch_part.current_stage = first_stage

#         batch_part.save()

#     # =====================================
#     # REDIRECT TO fill_stage
#     # =====================================

#     return redirect(

#         "fill_stage",

#         form_id=form_obj.id,

#         stage_id=first_stage.id,

#         part_id=batch_part.id

#     )

@login_required
def start_workorder_form(

    request,

    form_id,

    item_id

):

    try:

        import uuid

        # =====================================
        # GET FORM
        # =====================================

        form_obj = get_object_or_404(

            Form,

            id=form_id

        )

        # =====================================
        # GET WORKORDER ITEM
        # =====================================

        item = get_object_or_404(

            WorkOrderItem,

            id=item_id

        )

        # =====================================
        # GET FIRST STAGE
        # =====================================

        first_stage = (

            form_obj.stages

            .order_by("order")

            .first()

        )

        # =====================================
        # NO STAGE FOUND
        # =====================================

        if not first_stage:

            messages.error(

                request,

                "No stages found for this form."

            )

            return redirect(

                "workorder_kanban",

                wo_id=item.work_order.id

            )

        # =====================================
        # PART DISPLAY
        # =====================================

        part_display = item.product_code

        # =====================================
        # CREATE / GET FORM BATCH
        # =====================================

        form_batch, created = (

            FormBatch.objects.get_or_create(

                form=form_obj,

                batch_id=item.work_order.wo_number,

                defaults={

                    "company_name":
                        item.work_order.customer_name,

                    "rfq_ref_id":
                        item.work_order.rfq_ref,

                    "flow_id":
                        uuid.uuid4(),

                    "created_by":
                        request.user,

                    "current_stage":
                        first_stage

                }

            )

        )

        # =====================================
        # UPDATE COMPANY NAME
        # =====================================

        form_batch.company_name = (
            item.work_order.customer_name
        )
        form_batch.rfq_ref_id = (
            item.work_order.rfq_ref
        )
        # =====================================
        # DO NOT RESET CURRENT STAGE
        # =====================================

        if not form_batch.current_stage:

            form_batch.current_stage = first_stage

        form_batch.save()

        # =====================================
        # GET EXISTING BATCH PART
        # =====================================

        batch_part = BatchPart.objects.filter(

            batch=form_batch,

            part_id=part_display

        ).first()

        # =====================================
        # CREATE NEW PART
        # =====================================

        if not batch_part:

            batch_part = BatchPart.objects.create(

                batch=form_batch,

                part_id=part_display,

                current_stage=first_stage

            )

        # =====================================
        # UPDATE EXISTING PART
        # =====================================

        else:

            batch_part.batch = form_batch

            if not batch_part.current_stage:

                batch_part.current_stage = first_stage

            batch_part.save()

        # =====================================
        # FINAL SAFETY CHECK
        # =====================================

        if not batch_part.current_stage:

            batch_part.current_stage = first_stage

            batch_part.save()

        # =====================================
        # REDIRECT
        # =====================================

        return redirect(

            "fill_stage",

            form_id=form_obj.id,

            stage_id=batch_part.current_stage.id,

            part_id=batch_part.id

        )

    except Exception as e:

        import traceback

        return HttpResponse(

            f"<h1>ERROR:</h1>"
            f"<pre>{str(e)}</pre>"
            f"<hr>"
            f"<pre>{traceback.format_exc()}</pre>"

        )